"""
Parseur manifest navire à grue (format chinois Excel) → Pré-masque IPAKI.

Source : Excel multilingue (FR/EN/ZH) — 1 ligne = 1 B/L avec colonne DESCRIPTION
         contenant type véhicule, MODEL, N° VIN (un par ligne), ENGINE NO, YEAR, transit.
Sortie  : DataFrame 1 ligne par VIN/châssis, colonnes pré-masque IPAKI.
"""
from __future__ import annotations

import io
import re
from typing import Optional

import pandas as pd

# ---------------------------------------------------------------------------
# Regex utilitaires
# ---------------------------------------------------------------------------
# Marqueurs YEAR (nombreuses variantes trouvées sur données réelles)
_YEAR_RE = re.compile(
    r"(?:YEAR\s+OF\s+MANUFACTURE|MANUFACTURE\s+YEAR|MADE\s+OF\s+YEAR"
    r"|MADE\s+YEAR|DATE\s+OF\s+MANUFACTURE|YEAR|ANNEE?)[^0-9]*"
    r"(\d{4})",
    re.I,
)

# MODEL (colonnes : et ： full-width) — capture la valeur sans ponctuation initiale
_MODEL_RE = re.compile(r"MODEL(?:\s+NO)?[^\w\n]*[:：]\s*([A-Z0-9][A-Z0-9_\-\.]{2,})", re.I)

# VIN / CHASSIS sur même ligne que le label
_VIN_INLINE_RE = re.compile(
    r"(?:VIN\s+NO|CHASSIS\s+NUMBER|VIN)[^\w\n]*[:：.]\s*([A-Z0-9][A-Z0-9/\-]{8,})",
    re.I,
)

# Ligne contenant un label VIN/ENGINE (avec ou sans ENGINE NO en colonne)
# Accepte "VIN NO." seul ou "VIN NO.    ENGINE NO." (colonnes espacées)
_VIN_HEADER_RE = re.compile(
    r"^\s*(?:VIN(?:\s*(?:NO|\/ENGINE\s+NO)?)?|CHASSIS\s+NUMBER)"
    r"[\.：:\s]*(?:ENGINE\s+NO[\.：:]?)?\s*$",
    re.I,
)

# Token VIN "seul" : alphanumériques majuscules, longueur 10–22
_VIN_TOKEN_RE = re.compile(r"\b([A-Z0-9]{10,22})\b")

# Destinations transit
_TRANSIT_RE = re.compile(
    r"(?:TRANSIT\s+TO|CARGO\s+IN\s+TRANSIT\s+TO|FINAL\s+DESTINATION\s*[:：]"
    r"|DESTINATION\s+FINALE?\s*[:：]|ABIDJAN\s+TRANSIT\s+TO)\s*([A-Z][A-Z\s,À-Ü]+)",
    re.I,
)

# NEW / USED dans la description
_NEW_RE = re.compile(r"\b(NEW|NEUF|NEUVE|NEUFS)\b", re.I)
_USED_RE = re.compile(r"\b(USED|USAGE[DS]?|OCCASION|OLD|ANCIEN)\b", re.I)

# BSC / COC référence
_BSC_RE = re.compile(r"BSC\s*[#N°.]*\s*([A-Z0-9\-]+)", re.I)
_COC_RE = re.compile(r"COC\s+NO[.：:]\s*([^\s\n]+)", re.I)

# Mots-clés qui délimitent un nouveau groupe de véhicule dans la description
_GROUP_BOUNDARY_RE = re.compile(
    r"^\s*(?:\d+\s+)?(?:UNITS?\s+(?:OF\s+)?)?([A-Z][A-Z0-9 /\-&]*"
    r"(?:TRUCK|TRAILER|TRACTOR|BUS|VAN|EXCAVATOR|LOADER|CRANE|MIXER|TANKER"
    r"|VEHICLE|CAMION|DUMP|FENCE|SPRINKLER|TIPPER|FLATBED|REFRIGERATOR|AERIAL))\s*$",
    re.I,
)

# Marques connues (pour extraction MARQUE)
_BRANDS = [
    "SINOTRUK", "SHACMAN", "FOTON", "HOWO", "VOLVO", "LOVOL",
    "CATERPILLAR", "XCMG", "SANY", "LIUGONG", "XGMA", "SUNLION",
    "SUNFLOWER", "ZHONGTONG", "YUTONG", "DONGFENG", "JAC", "ISUZU",
    "MERCEDES", "MAN", "SCANIA", "DAF", "RENAULT", "HOMAN",
]

# Marques extraites du premier modèle connu de la description
def _extract_brand(desc_text: str, model: str) -> str:
    """Tente d'extraire la marque depuis le texte ou le préfixe du modèle."""
    text_up = desc_text.upper()
    for brand in _BRANDS:
        if brand in text_up:
            return brand
    # Fallback : premières lettres alphabétiques du modèle (ex. LZZ → inconnu, SX → SHACMAN)
    _model_brands = {
        "ZZ": "SINOTRUK", "LZZ": "SINOTRUK", "HLH": "SINOTRUK",
        "SX": "SHACMAN", "LZGJ": "SHACMAN",
        "LA9": "DIVERS", "LFY": "DIVERS",
        "YV2": "VOLVO",
        "LDY": "ZHONGTONG", "LCKS": "ZHONGTONG",
        "LEZ": "HOMAN", "LEZA": "HOMAN",
        "LCK": "ZHONGTONG",
        "DFH": "DONGFENG",
        "FTC": "LOVOL",
    }
    m = model.upper() if model else ""
    for prefix, brand in _model_brands.items():
        if m.startswith(prefix):
            return brand
    # Tentative depuis le VIN
    return ""


# ---------------------------------------------------------------------------
# Détection de la ligne d'en-tête dans le fichier source
# ---------------------------------------------------------------------------
_SOURCE_ALIASES = {
    "bl":          ["b/l no", "b/l", "bl no", "提单号码"],
    "marks":       ["marks", "唛", "mark"],
    "qty":         ["no. of package", "no of package", "件数", "no.", "qty"],
    "desc":        ["description", "货", "货名"],
    "weight":      ["g.weight", "weight", "毛重", "poids"],
    "volume":      ["measurement", "cbm", "尺码", "volume"],
    "shipper":     ["shipper", "货主", "expéditeur"],
    "consignee":   ["consignee", "收货人", "destinataire"],
    "notify":      ["notify", "通知人"],
    "term":        ["term", "备注"],
}


def _find_header(raw: pd.DataFrame) -> int:
    all_aliases = [a for aliases in _SOURCE_ALIASES.values() for a in aliases]
    best_row, best_score = 0, 0
    for i in range(min(10, len(raw))):
        row_text = " ".join(str(v).lower() for v in raw.iloc[i] if pd.notna(v))
        score = sum(1 for a in all_aliases if a in row_text)
        if score > best_score:
            best_score, best_row = score, i
    return best_row


def _find_col_idx(columns: list, key: str) -> Optional[int]:
    aliases = _SOURCE_ALIASES.get(key, [key])
    for i, col in enumerate(columns):
        col_l = str(col).lower().strip()
        for a in aliases:
            if a in col_l:
                return i
    return None


# ---------------------------------------------------------------------------
# Parsing de la colonne DESCRIPTION → liste de véhicules
# ---------------------------------------------------------------------------

def _clean(s) -> str:
    """Normalise les espaces et caractères full-width courants."""
    if not isinstance(s, str):
        s = "" if (s is None or (isinstance(s, float) and pd.isna(s))) else str(s)
    # Full-width colon/equals → ASCII
    s = s.replace("：", ":").replace("＝", "=").replace("，", ",")
    return s.strip()


def _split_vin_engine_line(line: str) -> list[str]:
    """Extrait les tokens VIN d'une ligne type 'VIN ENG' ou 'VIN1/VIN2'.
    Retourne uniquement les tokens qui ressemblent à un N° châssis."""
    # Slash-séparés (ex. VIN1/VIN2 ou VIN/ENGINE avec /)
    tokens = re.split(r"[/\s]+", line.strip())
    vins = []
    for t in tokens:
        t = t.strip().upper()
        # Un VIN/châssis : alphanum, 10-22 chars, pas que des chiffres
        if re.match(r'^[A-Z0-9]{10,22}$', t) and not t.isdigit():
            # Exclure les modèles (contiennent souvent des lettres ET chiffres en alternance
            # mais les modèles sont généralement < 15 chars avec patterns spéciaux)
            vins.append(t)
    return vins


def _parse_description(desc: str, total_qty: int) -> list[dict]:
    """
    Parse la colonne DESCRIPTION et retourne une liste de dicts, un par véhicule :
    {
      'vehicle_type': str,
      'model': str,
      'year': str,
      'vin': str,           # peut être vide si non trouvé
      'engine': str,
      'brand': str,
      'is_new': bool | None,
      'transit_dest': str,
      'bsc': str,
      'coc': str,
    }
    """
    desc = _clean(desc)
    lines = desc.splitlines()

    # ── Extraction globale (valeurs communes à tout le BL) ─────────────────
    transit_m = _TRANSIT_RE.search(desc)
    if transit_m:
        _raw_dest = transit_m.group(1).strip()
        # Conserver uniquement le nom du pays/ville : stop avant les mots parasites
        _dest_stop = re.split(
            r"[,،،\n]|(?:\s+(?:VIA|AT|CARRIER|CARGO|TRANSIT|AT\s+THE|BY\s+CONSIGNEE))\b",
            _raw_dest, maxsplit=1, flags=re.I
        )
        transit_dest = _dest_stop[0].strip().rstrip("，,. ").strip()
        transit_dest = re.sub(r"\s+", " ", transit_dest).strip()
    else:
        transit_dest = ""

    is_new: Optional[bool] = None
    if _NEW_RE.search(desc):
        is_new = True
    if _USED_RE.search(desc):
        is_new = False

    bsc_m = _BSC_RE.search(desc)
    bsc = bsc_m.group(1).strip() if bsc_m else ""
    coc_m = _COC_RE.search(desc)
    coc = coc_m.group(1).strip() if coc_m else ""

    # ── Découpage en groupes de véhicules ───────────────────────────────────
    # Un groupe commence à chaque nouveau MODEL: ou après un type de véhicule
    groups: list[dict] = []

    def _new_group(vtype="") -> dict:
        return {
            "vehicle_type": vtype,
            "model": "",
            "year": "",
            "brand": "",
            "vins": [],
            "engines": [],
            "in_vin_block": False,  # flag: on lit des VINs ligne par ligne
            "vin_engine_paired": False,  # les VINs et ENGINEs sont sur la même ligne
        }

    cur = _new_group()
    in_vin_header = False  # True quand la ligne d'en-tête VIN/ENGINE vient d'être vue

    for line in lines:
        line_clean = _clean(line)
        line_up = line_clean.upper()

        if not line_clean:
            in_vin_header = False
            continue

        # ── MODEL ────────────────────────────────────────────────────────────
        m_model = _MODEL_RE.search(line_clean)
        if m_model:
            new_model = m_model.group(1).strip()
            if cur["model"] and new_model != cur["model"]:
                # Nouveau modèle = nouveau groupe de véhicule
                groups.append(cur)
                cur = _new_group(cur["vehicle_type"])
            cur["model"] = new_model
            in_vin_header = False
            continue

        # ── YEAR ─────────────────────────────────────────────────────────────
        m_year = _YEAR_RE.search(line_clean)
        if m_year and not cur["year"]:
            cur["year"] = m_year.group(1)
            in_vin_header = False
            cur["in_vin_block"] = False
            continue

        # ── VIN INLINE (label + valeur sur la même ligne) ───────────────────
        m_vin_inline = _VIN_INLINE_RE.search(line_clean)
        if m_vin_inline:
            raw_vins = m_vin_inline.group(1)
            # Peut contenir VIN1/VIN2/VIN3
            extracted = _split_vin_engine_line(raw_vins)
            cur["vins"].extend(extracted)
            in_vin_header = False
            cur["in_vin_block"] = False
            continue

        # ── EN-TÊTE VIN (label seul ou "VIN NO.    ENGINE NO." en colonnes) ──
        # Détecte aussi le format "VIN NO.                    ENGINE NO." avec espaces
        _vin_hdr_match = _VIN_HEADER_RE.match(line_clean)
        if not _vin_hdr_match:
            # Variante : ligne commençant par VIN NO et contenant ENGINE NO après des espaces
            _vin_hdr_match = re.match(
                r"^\s*VIN\s*NO[\.：:.]?\s{2,}ENGINE\s+NO[\.：:.]?\s*$", line_clean, re.I
            )
        if _vin_hdr_match:
            in_vin_header = True
            cur["in_vin_block"] = True
            cur["vin_engine_paired"] = bool(re.search(r"ENGINE", line_up))
            continue

        # ── LIGNE DE VIN (suite de l'en-tête VIN) ───────────────────────────
        if in_vin_header and cur["in_vin_block"]:
            # Ignorer les lignes qui ressemblent à d'autres labels
            if any(kw in line_up for kw in
                   ["YEAR", "MODEL", "ENGINE NO:", "VIN NO:", "TRANSIT", "UNIT",
                    "PACKAGE", "DECK", "CARGO", "CARRIER", "COST", "RISK"]):
                in_vin_header = False
                cur["in_vin_block"] = False
            else:
                # Extraire tous les tokens alphanumériques (VIN + éventuellement ENGINE)
                tokens = [t for t in re.split(r"[\s/]+", line_clean)
                          if re.match(r'^[A-Z0-9]{10,22}$', t.upper()) and not t.isdigit()]
                if tokens:
                    if cur["vin_engine_paired"] and len(tokens) >= 2:
                        # Premier token = VIN, deuxième = ENGINE
                        cur["vins"].append(tokens[0].upper())
                        cur["engines"].append(tokens[1].upper())
                    elif tokens:
                        cur["vins"].append(tokens[0].upper())
                        if len(tokens) > 1:
                            cur["engines"].append(tokens[1].upper())
                    continue
                # Si aucun token VIN sur cette ligne → fin du bloc VIN
                in_vin_header = False
                cur["in_vin_block"] = False

        # ── TYPE DE VÉHICULE (première ligne significative du groupe) ────────
        # Ignorer les lignes BSC/COC/TRANSIT/quantité
        skip_patterns = ["BSC", "COC", "TRANSIT", "CARGO IN", "CARRIER",
                         "FREIGHT", "ON DECK", "AT THE COST", "BY CONSIGNEE",
                         "ABIDJAN", "COTE D", "IVOIRE", "RISK AND"]
        if not cur["vehicle_type"] and not any(p in line_up for p in skip_patterns):
            # Ligne descriptive = type de véhicule (exclure les chiffres seuls)
            candidate = re.sub(r"^\d+\s+UNITS?\s+(?:OF\s+)?", "", line_clean, flags=re.I).strip()
            candidate = re.sub(r"^\d+\s+PACKAGES?.*$", "", candidate, flags=re.I).strip()
            candidate = re.sub(r"BSC#.*", "", candidate, flags=re.I).strip()
            if candidate and len(candidate) > 3:
                cur["vehicle_type"] = candidate

    # Dernier groupe
    if cur["model"] or cur["vins"] or cur["vehicle_type"]:
        groups.append(cur)

    # ── Calcul global YEAR (fallback si non trouvé dans un groupe) ──────────
    global_year_m = _YEAR_RE.search(desc)
    global_year = global_year_m.group(1) if global_year_m else ""

    # ── Construction des lignes de sortie ────────────────────────────────────
    result: list[dict] = []
    for g in groups:
        model = g["model"]
        year = g["year"] or global_year
        vtype = g["vehicle_type"]
        brand = _extract_brand(vtype + " " + desc[:300], model)
        vins = g["vins"] or [""]  # au moins une ligne même sans VIN
        engines = g["engines"]

        for i, vin in enumerate(vins):
            engine = engines[i] if i < len(engines) else ""
            result.append({
                "vehicle_type": vtype,
                "model": model,
                "year": year,
                "vin": vin,
                "engine": engine,
                "brand": brand,
                "is_new": is_new,
                "transit_dest": transit_dest,
                "bsc": bsc,
                "coc": coc,
            })

    # Si rien extrait, créer des lignes vides basées sur la quantité
    if not result:
        for _ in range(max(1, total_qty)):
            result.append({
                "vehicle_type": "",
                "model": "",
                "year": global_year,
                "vin": "",
                "engine": "",
                "brand": "",
                "is_new": is_new,
                "transit_dest": transit_dest,
                "bsc": bsc,
                "coc": coc,
            })

    return result


# ---------------------------------------------------------------------------
# Lecture du fichier source
# ---------------------------------------------------------------------------

def parse_crane_manifest(file_bytes: bytes, filename: str) -> pd.DataFrame:
    """Lit un manifest navire à grue Excel et retourne un DataFrame normalisé."""
    fname = filename.lower()
    try:
        if fname.endswith(".xlsx"):
            raw = pd.read_excel(io.BytesIO(file_bytes), header=None, engine="openpyxl")
        else:
            raw = pd.read_excel(io.BytesIO(file_bytes), header=None, engine="xlrd")
    except Exception as e:
        raise ValueError(f"Impossible de lire '{filename}' : {e}") from e

    if raw.empty:
        raise ValueError(f"Fichier vide : {filename}")

    # Infos de l'en-tête globale (navire, voyage, POL, POD)
    header_global = {}
    for i in range(min(5, len(raw))):
        row_text = " ".join(str(v) for v in raw.iloc[i] if pd.notna(v))
        m_ship = re.search(r"SHIP'?S?\s+NAME[^:]*:\s*(.+?)(?:\s{2,}|$)", row_text, re.I)
        if m_ship:
            # Ex : "21 Coral V.26140" → navire = "21 Coral", voyage = "V.26140"
            parts = m_ship.group(1).strip().split()
            # Le code voyage contient souvent un point ou une lettre suivie de chiffres
            voyage_tokens = [p for p in parts if re.search(r'[A-Z]\.\d+|\d{4,}', p, re.I)]
            header_global["voyage"] = voyage_tokens[-1] if voyage_tokens else parts[-1]
            header_global["navire"] = " ".join(parts[:-1] if voyage_tokens else parts)

        m_pol = re.search(r"LOADING\s+PORT\s*:\s*(.+?)(?:\s{2,}|$)", row_text, re.I)
        if m_pol:
            header_global["pol"] = m_pol.group(1).strip()

        m_pod = re.search(r"DISCHARGE\s+PORT[^:]*:\s*(.+?)(?:\s{2,}|$)", row_text, re.I)
        if m_pod:
            header_global["pod"] = m_pod.group(1).strip()

    # Détection de la ligne d'en-tête des données
    header_row = _find_header(raw)
    columns = [str(v).strip() if pd.notna(v) else f"_c{i}"
               for i, v in enumerate(raw.iloc[header_row])]
    df = raw.iloc[header_row + 1:].copy()
    df.columns = columns
    df = df.dropna(how="all").reset_index(drop=True)

    # Index des colonnes utiles
    col_bl = _find_col_idx(columns, "bl")
    col_qty = _find_col_idx(columns, "qty")
    col_desc = _find_col_idx(columns, "desc")
    col_weight = _find_col_idx(columns, "weight")
    col_volume = _find_col_idx(columns, "volume")
    col_consignee = _find_col_idx(columns, "consignee")
    col_shipper = _find_col_idx(columns, "shipper")

    def _get_cell(row, col_idx):
        if col_idx is None or col_idx >= len(row):
            return ""
        v = row.iloc[col_idx]
        return "" if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v).strip()

    rows_out = []
    nbre = 0

    for _, row in df.iterrows():
        bl = _get_cell(row, col_bl)
        if not bl or bl.lower() in ("nan", "total:", "total", ""):
            continue

        qty_raw = _get_cell(row, col_qty)
        try:
            total_qty = int(float(qty_raw)) if qty_raw else 1
        except (ValueError, TypeError):
            total_qty = 1

        desc = _get_cell(row, col_desc)
        total_weight_str = _get_cell(row, col_weight)
        total_volume_str = _get_cell(row, col_volume)
        consignee_raw = _get_cell(row, col_consignee)
        shipper_raw = _get_cell(row, col_shipper)

        # Poids / volume bruts — None si absent (évite les 0.0 fictifs dans l'export)
        try:
            _w = float(str(total_weight_str).replace(",", "").replace(" ", ""))
            total_weight = _w if _w > 0 else None
        except (ValueError, TypeError):
            total_weight = None
        try:
            _v = float(str(total_volume_str).replace(",", "").replace(" ", ""))
            total_volume = _v if _v > 0 else None
        except (ValueError, TypeError):
            total_volume = None

        # Client = première ligne du consignee (avant le retour à la ligne)
        client = consignee_raw.split("\n")[0].strip() if consignee_raw else ""

        # Parse la DESCRIPTION
        vehicles = _parse_description(desc, total_qty)
        n_vehicles = len(vehicles)

        for v in vehicles:
            nbre += 1

            # Poids et volume par unité — chercher d'abord dans la desc
            # Pattern : "1 UNIT/11520KGS/79.54CBM" ou "11520 KGS"
            unit_weight: float | None = None
            unit_volume: float | None = None
            w_in_desc = re.search(r"(\d[\d,\.]+)\s*KGS?", v["vehicle_type"] + desc[:500], re.I)
            vol_in_desc = re.search(r"(\d[\d\.]+)\s*CBM", v["vehicle_type"] + desc[:500], re.I)
            if w_in_desc and n_vehicles == 1:
                try:
                    _wv = float(w_in_desc.group(1).replace(",", ""))
                    if _wv > 0:
                        unit_weight = _wv
                except ValueError:
                    pass
            if vol_in_desc and n_vehicles == 1:
                try:
                    _vv = float(vol_in_desc.group(1).replace(",", ""))
                    if _vv > 0:
                        unit_volume = _vv
                except ValueError:
                    pass

            # Fallback : division équitable du total BL
            if unit_weight is None and total_weight is not None and n_vehicles > 0:
                unit_weight = round(total_weight / n_vehicles, 1)
            if unit_volume is None and total_volume is not None and n_vehicles > 0:
                unit_volume = round(total_volume / n_vehicles, 3)

            # TYPE/TAILLE : C < 15m³, V 15-50m³, T > 50m³
            # "" si volume inconnu (ne pas classifier arbitrairement en T)
            if unit_volume is not None and unit_volume > 0:
                if unit_volume < 15:
                    size_type = "C"
                elif unit_volume <= 50:
                    size_type = "V"
                else:
                    size_type = "T"
            else:
                size_type = ""  # volume inconnu → à compléter manuellement

            # ETAT
            if v["is_new"] is True:
                etat = "VEHICULES NEUFS, VOITURES NEUVES, CAMION"
            elif v["is_new"] is False:
                etat = "VEHICULES USAGES,VOITURES OCCASIONS"
            else:
                etat = ""

            # NATURE BL : Import si destination = POD, Transbo si autre pays
            pod_city = (header_global.get("pod") or "ABIDJAN").upper()
            transit = v["transit_dest"].upper()
            if transit and not any(p in transit for p in ["ABIDJAN", "COTE", "IVOIRE", "CI"]):
                nature_bl = "Transbo"
            else:
                nature_bl = "Import"

            # DESTINATION FINALE
            dest_finale = v["transit_dest"] if v["transit_dest"] else (header_global.get("pod") or "ABIDJAN")

            # MARQUE & MODELE
            marque = v["brand"]
            modele = v["model"]
            marque_modele = f"{marque} {modele}".strip() if marque else modele

            # OBSERVATION
            obs_parts = []
            if v["bsc"]:
                obs_parts.append(f"BSC#{v['bsc']}")
            if v["coc"]:
                obs_parts.append(f"COC#{v['coc']}")
            if v["transit_dest"]:
                obs_parts.append(f"TRANSIT VERS {v['transit_dest']}")
            observation = " | ".join(obs_parts)

            # Poids IPAKI en tonnes
            poids_ton = round(unit_weight / 1000, 3) if unit_weight else ""

            rows_out.append({
                "NBRE":                         nbre,
                "MODE DE TRANSPORT":            "",
                "ESCALE TETRAX":                "",
                "NATURE BL":                    nature_bl,
                "POL TETRAX":                   header_global.get("pol", ""),
                "POD TETRAX":                   pod_city,
                "FINAL DESTINATION TETRAX":     dest_finale,
                "POIDS TETRAX (KG)":            int(unit_weight) if unit_weight else "",
                "TYPE / TAILLE":                size_type,
                "VOLUME TETRAX":                unit_volume if unit_volume else "",
                "BL":                           bl,
                "ESCALE IPAKI":                 "",
                "POL IPAKI":                    "",
                "POD IPAKI":                    "",
                "FINAL DESTINATION IPAKI":      dest_finale,
                "MARQUE":                       marque,
                "MODELE":                       modele,
                "MARQUE & MODELE":              marque_modele,
                "ETAT":                         etat,
                "ANNEE DE FABRICATION":         v["year"],
                "CHÂSSIS":                      v["vin"],
                "TYPE D'ACTION":               "IMPORT" if nature_bl == "Import" else "TRANSBO",
                "POIDS IPAKI (TON)":            poids_ton,
                "VOLUME":                       unit_volume,
                "BLItem YardItemCode":          f"VEH {'< 15m3' if size_type == 'C' else '> 15m3'}",
                "CLIENT":                       client,
                "OBSERVATION":                  observation,
                # Colonnes d'aide (masquables)
                "_BL_SOURCE":                   bl,
                "_VEHICLE_TYPE":                v["vehicle_type"],
                "_ENGINE_NO":                   v["engine"],
                "_SHIPPER":                     shipper_raw.split("\n")[0].strip() if shipper_raw else "",
            })

    return pd.DataFrame(rows_out)


# ---------------------------------------------------------------------------
# Export Excel pré-masque
# ---------------------------------------------------------------------------

def generate_premasque_excel(df: pd.DataFrame, navire: str = "", voyage: str = "") -> bytes:
    """Génère un Excel au format pré-masque IPAKI depuis le DataFrame normalisé.

    - En-tête gris clair, colonnes ajustées
    - Colonnes d'aide (préfixées _) dans un onglet séparé
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Colonnes à exporter (sans les colonnes _)
    export_cols = [c for c in df.columns if not c.startswith("_")]
    df_export = df[export_cols]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pré-Masque IPAKI"

    # Style en-tête
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF", size=9)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # En-têtes
    for col_idx, col_name in enumerate(export_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    # Données
    data_align = Alignment(vertical="center")
    for row_idx, (_, row) in enumerate(df_export.iterrows(), 2):
        for col_idx, col_name in enumerate(export_cols, 1):
            val = row[col_name]
            # Convertir numpy types
            if hasattr(val, "item"):
                val = val.item()
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = data_align
            cell.border = border
            # Coloration alternée par BL
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F2F2F2")

    # Largeurs de colonnes
    col_widths = {
        "NBRE": 6, "MODE DE TRANSPORT": 14, "ESCALE TETRAX": 14,
        "NATURE BL": 10, "POL TETRAX": 12, "POD TETRAX": 12,
        "FINAL DESTINATION TETRAX": 18, "POIDS TETRAX (KG)": 14,
        "TYPE / TAILLE": 10, "VOLUME TETRAX": 12, "BL": 20,
        "ESCALE IPAKI": 14, "POL IPAKI": 12, "POD IPAKI": 12,
        "FINAL DESTINATION IPAKI": 18, "MARQUE": 14, "MODELE": 16,
        "MARQUE & MODELE": 22, "ETAT": 28, "ANNEE DE FABRICATION": 12,
        "CHÂSSIS": 20, "TYPE D'ACTION": 10, "POIDS IPAKI (TON)": 14,
        "VOLUME": 12, "BLItem YardItemCode": 16, "CLIENT": 28,
        "OBSERVATION": 30,
    }
    for col_idx, col_name in enumerate(export_cols, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 14)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 40

    # Onglet aide (colonnes _)
    helper_cols = [c for c in df.columns if c.startswith("_")]
    if helper_cols:
        ws2 = wb.create_sheet("Données brutes")
        helper_hdr_fill = PatternFill("solid", fgColor="70AD47")
        for col_idx, col_name in enumerate(helper_cols, 1):
            cell = ws2.cell(row=1, column=col_idx, value=col_name.lstrip("_"))
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = helper_hdr_fill
        for row_idx, (_, row) in enumerate(df[helper_cols].iterrows(), 2):
            for col_idx, col_name in enumerate(helper_cols, 1):
                val = row[col_name]
                if hasattr(val, "item"):
                    val = val.item()
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col_idx, col_name in enumerate(helper_cols, 1):
            ws2.column_dimensions[get_column_letter(col_idx)].width = 25
            ws2.row_dimensions[1].height = 25

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
