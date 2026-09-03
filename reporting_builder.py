"""
reporting_builder.py — Onglet Reporting : construction de la liste
prévisionnelle définitive à partir des manifestes déjà structurés, et
rapprochements avec la 1ère liste provisoire (service Reporting) et le
Discharging Container Summary.

Contexte (voir claude/ANALYSE_ONGLET_REPORTING_2026-09-03.md dans le projet
Claude) : les manifestes PDF font foi pour la liste prévisionnelle définitive.
Le Discharging Container Summary ne contient AUCUN numéro de B/L — le
rapprochement avec ce fichier se fait uniquement par numéro de conteneur, et
uniquement pour la catégorie Conteneur. Des colonnes du gabarit Reporting
(Agent, STATUTS, REMARQUES, ARRIVAL, CLIENT distinct du destinataire...) sont
des champs de booking absents du manifeste brut : elles restent vides dans la
liste générée, à charge du service Reporting de les compléter.

Aucune dépendance nouvelle : réutilise tracking.py (lecture des traitements
archivés + téléchargement des exports Excel) et manifest_parser.write_sheet
(mise en forme Excel identique au reste de l'application).
"""
import io
import re

import pandas as pd
import pdfplumber
from openpyxl import Workbook, load_workbook

import tracking
from manifest_parser import write_sheet


# ---------------------------------------------------------------------------
# Normalisation des clés de rapprochement
# ---------------------------------------------------------------------------
def normalize_bl(value) -> str:
    """Normalise un numéro de B/L pour comparaison entre sources hétérogènes :
    manifeste ('S329891233'), liste prévisionnelle onglet CONTENEUR (format
    identique) ou onglets RORO/BB ('S3-29891233', tiret après S3 — confirmé
    équivalent au format manifeste une fois le tiret retiré). Retire tout
    caractère non alphanumérique et met en majuscule."""
    if value is None:
        return ""
    s = str(value).strip().upper()
    if not s or s in ("NAN", "NONE"):
        return ""
    return re.sub(r"[^A-Z0-9]", "", s)


def normalize_container(value) -> str:
    """Normalise un numéro de conteneur (manifeste : 'CN:GCNU1400426' déjà
    concaténé ; Discharging Summary : préfixe 'GCNU' (colonne Ref.) + numéro
    '1400426' (colonne ID no.) en deux colonnes séparées) — retire
    espaces/tirets, met en majuscule."""
    if value is None:
        return ""
    s = str(value).strip().upper()
    if not s or s in ("NAN", "NONE"):
        return ""
    return re.sub(r"[^A-Z0-9]", "", s)


def _col(df: pd.DataFrame, name: str, default=""):
    """Retourne df[name] si la colonne existe (valeurs vides remplacées par
    default), sinon une série de `default` de la bonne longueur — évite les
    KeyError si une feuille source ne contient pas toutes les colonnes
    attendues (ex. manifeste conteneur-only, sans onglet Véhicule)."""
    if name in df.columns:
        return df[name].fillna(default)
    return pd.Series([default] * len(df), index=df.index)


# ---------------------------------------------------------------------------
# Étape 1 — Agrégation des manifestes déjà structurés pour un Navire/Voyage
# ---------------------------------------------------------------------------
# Depuis la fusion des onglets détail du 03/09 (manifest_parser.py,
# build_workbook_bytes), un classeur exporté ne contient plus 3 onglets
# détail séparés mais UN SEUL onglet "Détail Cargaison" (une ligne par
# unité — châssis/conteneur/colis — avec une colonne "Catégorie" parmi
# "Véhicules"/"Conteneurs"/"Colis", voir CATEGORY_LABELS et
# MERGED_DETAIL_COLUMNS dans manifest_parser.py). On lit cet onglet unique
# puis on scinde par catégorie.
DETAIL_SHEET_NAME = "Détail Cargaison"
CATEGORY_TO_KEY = {"Véhicules": "Vehicule", "Conteneurs": "Conteneur", "Colis": "Colis"}


def _read_sheet_as_df(wb, sheet_name: str) -> pd.DataFrame:
    """Lit une feuille openpyxl générée par write_sheet() (bandeau titre +
    ligne vide avant l'en-tête réel) et retourne un DataFrame propre. Repère
    la ligne d'en-tête dynamiquement (présence de 'BL_Numero' dans la ligne,
    quelle que soit sa position — l'onglet fusionné a 'Catégorie' en 1ère
    colonne, 'BL_Numero' en 2e) plutôt que de supposer un nombre fixe de
    lignes de bandeau ou une position de colonne fixe."""
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        if row and "BL_Numero" in row:
            header_idx = i
            break
    if header_idx is None:
        return pd.DataFrame()
    header = [str(c) if c is not None else "" for c in rows[header_idx]]
    data_rows = rows[header_idx + 1:]
    df = pd.DataFrame(data_rows, columns=header)
    df = df.dropna(how="all")
    return df


# Anciens onglets détail séparés (avant la fusion du 03/09 en un seul onglet
# "Détail Cargaison" avec colonne Catégorie) — conservés en repli pour que les
# exports archivés AVANT la fusion restent exploitables dans Reporting.
# Sans ce repli, un voyage dont TOUS les ports déjà traités l'ont été avant la
# fusion remonte "Aucune donnée détail trouvée" alors que les traitements
# existent bien (bug identifié 03/09, données archivées confirmées).
LEGACY_DETAIL_SHEETS = {
    "Vehicule":  "Détail Véhicules (par châssis)",
    "Conteneur": "Détail Conteneurs",
    "Colis":     "Détail Colis",
}
# Dans les anciens onglets, le poids par unité s'appelait "Poids_Kg" (déjà une
# ligne par unité, comme aujourd'hui) au lieu de "Poids_Unitaire_Kg".
LEGACY_WEIGHT_RENAME = {"Poids_Kg": "Poids_Unitaire_Kg"}


def _read_voyage_detail_from_workbook(wb) -> dict:
    """Retourne {cle_categorie: DataFrame} pour UN classeur archivé, en
    tolérant les deux structures rencontrées dans l'historique :
    - Nouvelle (depuis le 03/09) : un seul onglet "Détail Cargaison" avec une
      colonne "Catégorie" (Véhicules / Conteneurs / Colis).
    - Ancienne (exports archivés avant cette date) : 3 onglets séparés
      (voir LEGACY_DETAIL_SHEETS). Clés vides omises."""
    result = {}
    df_merged = _read_sheet_as_df(wb, DETAIL_SHEET_NAME)
    if not df_merged.empty and "Catégorie" in df_merged.columns:
        for label, key in CATEGORY_TO_KEY.items():
            df_cat = df_merged[df_merged["Catégorie"] == label]
            if not df_cat.empty:
                result[key] = df_cat
        if result:
            return result
    for key, sheet_name in LEGACY_DETAIL_SHEETS.items():
        df = _read_sheet_as_df(wb, sheet_name)
        if df.empty:
            continue
        df = df.rename(columns={k: v for k, v in LEGACY_WEIGHT_RENAME.items() if k in df.columns})
        result[key] = df
    return result


def fetch_voyage_detail(navire: str, voyage: str):
    """Agrège, pour un Navire/Voyage donné, l'onglet détail fusionné (voir
    DETAIL_SHEET_NAME) de TOUS les traitements déjà archivés pour ce couple
    (généralement un traitement par port de chargement) — seule source
    contenant la donnée ligne à ligne, la base de suivi ne stockant que des
    compteurs agrégés par traitement. Scinde ensuite par catégorie.

    Retourne (dict {catégorie: DataFrame}, DataFrame des traitements source
    effectivement utilisés, liste triée des ports de chargement couverts,
    diagnostic {total, sans_export, echec_telechargement, illisible, ok}) —
    le diagnostic sert à distinguer, côté page, "rien n'a été traité" de
    "des traitements existent mais leur export n'a pas pu être lu" (bug
    silencieux sinon : mêmes deux cas indiscernables à l'affichage)."""
    keys = list(CATEGORY_TO_KEY.values())
    diag = {"total": 0, "sans_export": 0, "echec_telechargement": 0, "illisible": 0, "ok": 0}
    log = tracking.read_log()
    if log.empty:
        return {k: pd.DataFrame() for k in keys}, pd.DataFrame(), [], diag
    all_sub = log[(log["navire"] == navire) & (log["voyage"] == voyage)].copy()
    diag["total"] = len(all_sub)
    sub = all_sub[all_sub["export_path"].notna()]
    diag["sans_export"] = diag["total"] - len(sub)
    if sub.empty:
        return {k: pd.DataFrame() for k in keys}, pd.DataFrame(), [], diag

    collected = {k: [] for k in keys}
    ports = set()
    used_rows = []
    for _, row in sub.iterrows():
        data = tracking.get_archive_file(row["export_path"])
        if not data:
            diag["echec_telechargement"] += 1
            continue
        try:
            wb = load_workbook(io.BytesIO(data), data_only=True)
        except Exception:
            diag["illisible"] += 1
            continue
        by_cat = _read_voyage_detail_from_workbook(wb)
        if not by_cat:
            diag["illisible"] += 1
            continue
        any_found = False
        for key, df_cat in by_cat.items():
            if "Port_Chargement" in df_cat.columns:
                ports.update(p for p in df_cat["Port_Chargement"].dropna().unique() if p)
            collected[key].append(df_cat)
            any_found = True
        if any_found:
            used_rows.append(row)
            diag["ok"] += 1
        else:
            diag["illisible"] += 1

    result = {}
    for cat, parts in collected.items():
        result[cat] = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()
    used_df = pd.DataFrame(used_rows) if used_rows else pd.DataFrame()
    return result, used_df, sorted(ports), diag


def list_voyages_disponibles() -> pd.DataFrame:
    """Retourne les couples Navire/Voyage déjà traités (au moins un
    traitement archivé avec export), avec le nombre de traitements ET la
    liste des ports de chargement déjà couverts (plus parlant que le seul
    nombre de traitements — un port peut être retraité plusieurs fois) —
    pour peupler le sélecteur de la page Reporting sans que l'agent ait à
    ressaisir le nom du voyage.

    Relit les exports archivés de chaque voyage pour connaître ses ports
    (même logique que fetch_voyage_detail) — acceptable au volume actuel ;
    à mettre en cache (st.cache_data côté page) si le nombre de voyages
    archivés devient important."""
    log = tracking.read_log()
    cols = ["navire", "voyage", "nb_traitements", "derniere_maj", "ports"]
    if log.empty:
        return pd.DataFrame(columns=cols)
    sub = log[
        log["navire"].notna() & (log["navire"] != "") & (log["navire"] != "(navire non détecté)")
        & log["voyage"].notna() & (log["voyage"] != "")
    ]
    if sub.empty:
        return pd.DataFrame(columns=cols)
    grouped = (
        sub.groupby(["navire", "voyage"])
        .agg(nb_traitements=("id", "count"), derniere_maj=("horodatage", "max"))
        .reset_index()
        .sort_values("derniere_maj", ascending=False)
    )
    ports_col = []
    for _, r in grouped.iterrows():
        _, _, ports, _diag = fetch_voyage_detail(r["navire"], r["voyage"])
        ports_col.append(", ".join(ports) if ports else "aucun port exploitable")
    grouped["ports"] = ports_col
    return grouped


# ---------------------------------------------------------------------------
# Gabarit Reporting (RORO / CONTENEUR / BB) — colonnes retenues
# ---------------------------------------------------------------------------
# Champs marqués (booking) n'existent pas dans le manifeste PDF brut : ils
# restent vides dans la liste générée, à compléter par le service Reporting.
RORO_TEMPLATE_COLUMNS = [
    "Vessel", "Voyage", "Shipment#", "Agent",              # Agent = booking
    "POL", "POD", "Size", "Type",
    "Commodity/Model", "Model", "Weight(ton)", "CBM", "Equipment#",
    "STATUTS", "REMARQUES", "ARRIVAL",                      # booking
    "Etat", "Pays_Transit", "Nature_BL", "Chargeur_Nom", "Destinataire_Nom",  # bonus manifeste
]
CONTENEUR_TEMPLATE_COLUMNS = [
    "Vessel", "Voyage", "Shipment#", "POL", "POD",
    "Size", "Type", "Commodity/Model", "CLIENT", "Weight(ton)",
    "Equipment#", "Seal#", "Teus",
    "STATUTS", "REMARQUES", "ARRIVAL",                      # booking
    "Pays_Transit", "Nature_BL", "Chargeur_Nom",             # bonus manifeste
]
BB_TEMPLATE_COLUMNS = [
    "Vessel", "Voyage", "Shipment#", "Agent",               # Agent = booking
    "POL", "POD", "Type", "Commodity/Model", "Weight(ton)", "CBM",
    "Consignee", "Shipper",
    "STATUTS", "REMARQUES", "ARRIVAL",                      # booking
    "Pays_Transit", "Nature_BL",                             # bonus manifeste
]


def _to_ton(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce") / 1000.0


def build_liste_previsionnelle(dfs: dict) -> dict:
    """Convertit les feuilles détail agrégées (Vehicule/Conteneur/Colis) vers
    le gabarit Reporting (RORO/CONTENEUR/BB). Ne remplit que les champs
    physiquement présents dans le manifeste — voir le bandeau en tête de
    fichier. Ajoute une colonne technique _BL_norm (et _CONT_norm pour
    CONTENEUR) utilisée par les fonctions de rapprochement, à retirer avant
    tout export destiné à un agent."""
    out = {}

    df_v = dfs.get("Vehicule", pd.DataFrame())
    roro = pd.DataFrame({
        "Vessel": _col(df_v, "Navire"),
        "Voyage": _col(df_v, "Voyage"),
        "Shipment#": _col(df_v, "BL_Numero"),
        "Agent": "",
        "POL": _col(df_v, "Port_Chargement"),
        "POD": _col(df_v, "Port_Dechargement"),
        "Size": "",
        "Type": "RO",
        "Commodity/Model": (_col(df_v, "Marque") + " " + _col(df_v, "Modele")).str.strip(),
        "Model": _col(df_v, "Modele"),
        "Weight(ton)": _to_ton(_col(df_v, "Poids_Unitaire_Kg")),
        "CBM": "",
        "Equipment#": _col(df_v, "Chassis"),
        "STATUTS": "", "REMARQUES": "", "ARRIVAL": "",
        "Etat": _col(df_v, "Etat"),
        "Pays_Transit": _col(df_v, "Pays_Transit"),
        "Nature_BL": _col(df_v, "Nature_BL"),
        "Chargeur_Nom": _col(df_v, "Chargeur_Nom"),
        "Destinataire_Nom": _col(df_v, "Destinataire_Nom"),
    })
    roro["_BL_norm"] = roro["Shipment#"].map(normalize_bl)
    out["RORO"] = roro[RORO_TEMPLATE_COLUMNS + ["_BL_norm"]]

    df_c = dfs.get("Conteneur", pd.DataFrame())
    cont = pd.DataFrame({
        "Vessel": _col(df_c, "Navire"),
        "Voyage": _col(df_c, "Voyage"),
        "Shipment#": _col(df_c, "BL_Numero"),
        "POL": _col(df_c, "Port_Chargement"),
        "POD": _col(df_c, "Port_Dechargement"),
        "Size": "",
        "Type": _col(df_c, "Type_Colis"),
        "Commodity/Model": "",
        "CLIENT": _col(df_c, "Destinataire_Nom"),
        "Weight(ton)": _to_ton(_col(df_c, "Poids_Unitaire_Kg")),
        "Equipment#": _col(df_c, "No_Conteneur"),
        "Seal#": _col(df_c, "No_Scelle"),
        "Teus": "",
        "STATUTS": "", "REMARQUES": "", "ARRIVAL": "",
        "Pays_Transit": _col(df_c, "Pays_Transit"),
        "Nature_BL": _col(df_c, "Nature_BL"),
        "Chargeur_Nom": _col(df_c, "Chargeur_Nom"),
    })
    cont["_BL_norm"] = cont["Shipment#"].map(normalize_bl)
    cont["_CONT_norm"] = cont["Equipment#"].map(normalize_container)
    out["CONTENEUR"] = cont[CONTENEUR_TEMPLATE_COLUMNS + ["_BL_norm", "_CONT_norm"]]

    df_d = dfs.get("Colis", pd.DataFrame())
    bb = pd.DataFrame({
        "Vessel": _col(df_d, "Navire"),
        "Voyage": _col(df_d, "Voyage"),
        "Shipment#": _col(df_d, "BL_Numero"),
        "Agent": "",
        "POL": _col(df_d, "Port_Chargement"),
        "POD": _col(df_d, "Port_Dechargement"),
        "Type": _col(df_d, "Type_Colis"),
        "Commodity/Model": "",
        "Weight(ton)": _to_ton(_col(df_d, "Poids_Unitaire_Kg")),
        "CBM": _col(df_d, "Volume_CBM"),
        "Consignee": _col(df_d, "Destinataire_Nom"),
        "Shipper": _col(df_d, "Chargeur_Nom"),
        "STATUTS": "", "REMARQUES": "", "ARRIVAL": "",
        "Pays_Transit": _col(df_d, "Pays_Transit"),
        "Nature_BL": _col(df_d, "Nature_BL"),
    })
    bb["_BL_norm"] = bb["Shipment#"].map(normalize_bl)
    out["BB"] = bb[BB_TEMPLATE_COLUMNS + ["_BL_norm"]]

    return out


def build_previsionnelle_workbook_bytes(previs: dict, navire: str, voyage: str) -> io.BytesIO:
    """Classeur Excel (3 onglets RORO/CONTENEUR/BB) prêt à être vérifié/
    complété par le service Reporting — même mise en forme que le reste de
    l'application (write_sheet)."""
    title = [
        f"Navire : {navire}", f"Voyage : {voyage}",
        "Liste prévisionnelle générée depuis les manifestes structurés — "
        "à vérifier, ajuster et compléter (Agent/STATUTS/REMARQUES/ARRIVAL/CLIENT...)",
    ]
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_key, cols in (
        ("RORO", RORO_TEMPLATE_COLUMNS),
        ("CONTENEUR", CONTENEUR_TEMPLATE_COLUMNS),
        ("BB", BB_TEMPLATE_COLUMNS),
    ):
        df = previs.get(sheet_key, pd.DataFrame())
        visible = [c for c in cols if c in df.columns]
        data = df[visible].reset_index(drop=True) if not df.empty else pd.DataFrame(columns=visible)
        ws = wb.create_sheet(sheet_key)
        write_sheet(ws, data, title_lines=title)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Étape 2 — Rapprochement avec la 1ère liste provisoire (Reporting)
# ---------------------------------------------------------------------------
# Variantes de libellé tolérées pour les colonnes clés du fichier Reporting —
# un rapprochement basé sur une colonne non identifiée (ex. "Shipment #" au
# lieu de "Shipment#") produisait silencieusement un _BL_norm vide pour
# TOUTES les lignes, donc un rapprochement totalement faux (100% des B/L du
# manifeste signalés à tort comme "absents") sans aucun avertissement (bug
# silencieux corrigé le 03/09, retour utilisateur : "le matching doit être
# suffisamment robuste pour qu'on fasse confiance aux écarts affichés").
_BL_COLUMN_CANDIDATES = [
    "shipment#", "shipment #", "shipment", "shipment no", "shipment no.",
    "shipmentno", "shipment_no", "bl", "b/l", "bl#", "bl #", "bl no",
    "bl no.", "numero bl", "numéro bl", "n° bl", "no bl",
]
_CONT_COLUMN_CANDIDATES = [
    "equipment#", "equipment #", "equipment", "equipment no", "equipment no.",
    "container#", "container #", "container", "container no", "container no.",
    "no conteneur", "numero conteneur", "numéro conteneur", "n° conteneur",
]


def _find_column(df: pd.DataFrame, candidates: list):
    """Retourne le nom de colonne réel de df correspondant à l'un des
    libellés candidats (comparaison insensible à la casse et aux espaces
    superflus), ou None si aucune ne correspond — pour tolérer les variantes
    de libellé d'un fichier Reporting à l'autre sans jamais deviner
    silencieusement quelle colonne utiliser."""
    norm_map = {re.sub(r"\s+", " ", str(c)).strip().lower(): c for c in df.columns}
    for cand in candidates:
        if cand in norm_map:
            return norm_map[cand]
    return None


def parse_liste_provisoire(file_bytes: bytes, filename: str):
    """Lit un fichier liste prévisionnelle (format Reporting, .xls ou .xlsx,
    onglets RORO/CONTENEUR/BB) et retourne (dict {onglet: DataFrame} avec une
    colonne technique _BL_norm ajoutée (B/L normalisé), dict {onglet:
    message} pour tout onglet non vide où la colonne B/L n'a pas pu être
    identifiée — le rapprochement pour cet onglet est alors invalidé côté
    page plutôt que silencieusement faux). Noms d'onglet et de colonne
    reconnus avec tolérance à la casse/aux espaces (voir _find_column)."""
    engine = "xlrd" if filename.lower().endswith(".xls") else "openpyxl"
    out = {}
    warnings = {}
    xls = pd.ExcelFile(io.BytesIO(file_bytes), engine=engine)
    sheet_map = {re.sub(r"\s+", "", str(s)).strip().upper(): s for s in xls.sheet_names}
    for sheet in ("RORO", "CONTENEUR", "BB"):
        real_name = sheet_map.get(sheet)
        if real_name is None:
            out[sheet] = pd.DataFrame()
            continue
        df = xls.parse(real_name)
        bl_col = _find_column(df, _BL_COLUMN_CANDIDATES)
        if bl_col is not None:
            df["_BL_norm"] = df[bl_col].map(normalize_bl)
        else:
            if not df.empty:
                cols_apercu = ", ".join(str(c) for c in df.columns[:12])
                warnings[sheet] = (
                    f"colonne B/L introuvable dans l'onglet « {real_name} » "
                    f"(attendu : « Shipment# » ou équivalent — colonnes "
                    f"trouvées : {cols_apercu}{'…' if len(df.columns) > 12 else ''})"
                )
            df["_BL_norm"] = ""
        if sheet == "CONTENEUR":
            cont_col = _find_column(df, _CONT_COLUMN_CANDIDATES)
            if cont_col is not None:
                df["_CONT_norm"] = df[cont_col].map(normalize_container)
        out[sheet] = df
    return out, warnings


def reconcile_bl(df_manifeste: pd.DataFrame, df_provisoire: pd.DataFrame):
    """Rapproche deux jeux de B/L (colonne _BL_norm dans les deux, déjà
    normalisée). Retourne (B/L manifeste absents de la liste provisoire —
    'à ajouter', B/L liste provisoire absents des manifestes déjà traités —
    'à vérifier', nombre de B/L en commun)."""
    bl_manifeste = set(df_manifeste["_BL_norm"]) - {""} if "_BL_norm" in df_manifeste.columns else set()
    bl_provisoire = set(df_provisoire["_BL_norm"]) - {""} if "_BL_norm" in df_provisoire.columns else set()

    manquants = bl_manifeste - bl_provisoire
    en_trop = bl_provisoire - bl_manifeste
    communs = bl_manifeste & bl_provisoire

    df_manquants = (
        df_manifeste[df_manifeste["_BL_norm"].isin(manquants)].copy() if manquants else df_manifeste.iloc[0:0].copy()
    )
    df_en_trop = (
        df_provisoire[df_provisoire["_BL_norm"].isin(en_trop)].copy() if en_trop else df_provisoire.iloc[0:0].copy()
    )
    return df_manquants, df_en_trop, len(communs)


# ---------------------------------------------------------------------------
# Étape 3 — Rapprochement avec le Discharging Container Summary
# ---------------------------------------------------------------------------
# Format tabulaire à colonnes variables (IMDG/Ref./D.Berth souvent vides) :
# "Cell IMDG Ref. ID no. Type Kilogr. VGM Dir POD D.Berth FIN POL MV"
# Ex. de ligne : "142C502 ACLU 2788457 22DC 29271 Yes CIABJ CIABJ BEANR F"
#   (IMDG, Dir et D.Berth vides sur cette ligne)
# Ref. = préfixe conteneur (4 lettres, ex. ACLU/GCNU), ID no. = 5-8 chiffres
# → No_Conteneur = Ref.+ID no. concaténés, comparable au 'CN:ACLU2788457' du
# manifeste (voir normalize_container).
DISCHARGE_ROW_RE = re.compile(
    r"^(?P<cell>\S+)\s+"
    r"(?:(?P<imdg>\S+)\s+)?"
    r"(?P<ref>[A-Z]{4})\s+"
    r"(?P<idno>\d{5,8})\s+"
    r"(?P<type>\d{2}[A-Z]{2})\s+"
    r"(?P<kg>[\d,]+(?:\.\d+)?)\s+"
    r"(?P<vgm>Yes|No)\s+"
    r"(?P<rest>.+?)\s+"
    r"(?P<pol>[A-Z]{3,5})\s+"
    r"(?P<mv>[A-Z])$"
)


def parse_discharging_summary_pdf(file_bytes: bytes) -> pd.DataFrame:
    """Parse le PDF 'Discharging Container Summary' (tableau conteneur par
    conteneur, SANS numéro de B/L). Colonnes retournées : Cell, IMDG, Ref,
    ID_no, No_Conteneur (concaténé, normalisé — clé de rapprochement), Type,
    Kilogr, VGM, POD, FIN, POL, MV. Les lignes qui ne correspondent pas au
    format attendu (bandeaux, totaux, en-têtes) sont silencieusement
    ignorées plutôt que de lever une erreur — comportement volontaire, ce
    fichier contient beaucoup de lignes non-data (résumé par baie, totaux)."""
    rows = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.split("\n"):
                line = line.strip()
                m = DISCHARGE_ROW_RE.match(line)
                if not m:
                    continue
                gd = m.groupdict()
                rest_tokens = gd["rest"].split()
                pod = rest_tokens[0] if rest_tokens else ""
                fin = rest_tokens[-1] if rest_tokens else ""
                rows.append({
                    "Cell": gd["cell"],
                    "IMDG": gd["imdg"] or "",
                    "Ref": gd["ref"],
                    "ID_no": gd["idno"],
                    "No_Conteneur": normalize_container(gd["ref"] + gd["idno"]),
                    "Type": gd["type"],
                    "Kilogr": pd.to_numeric(gd["kg"].replace(",", ""), errors="coerce"),
                    "VGM": gd["vgm"],
                    "POD": pod,
                    "FIN": fin,
                    "POL": gd["pol"],
                    "MV": gd["mv"],
                })
    return pd.DataFrame(rows)


def reconcile_containers(df_manifeste_cont: pd.DataFrame, df_discharge: pd.DataFrame):
    """Rapproche les conteneurs de l'onglet CONTENEUR de la liste
    prévisionnelle générée (colonne _CONT_norm) avec le Discharging Container
    Summary (colonne No_Conteneur). Retourne (conteneurs manifeste absents du
    discharging summary, conteneurs discharging summary absents du
    manifeste, DataFrame des écarts de poids pour les conteneurs présents
    dans les deux — TOUT écart est remonté, sans seuil : décision explicite
    de l'utilisateur, le tri par écart absolu décroissant permet de prioriser
    visuellement, dup_warning : message si des numéros de conteneur sont
    dupliqués d'un côté ou de l'autre — un merge pd.merge() sur des clés
    dupliquées génère silencieusement un produit cartésien (ex. 2 lignes
    manifeste + 2 lignes discharge pour le même conteneur → 4 lignes
    fusionnées au lieu de 2, et le compteur 'conteneurs communs rapprochés'
    devient faux) ; on déduplique donc chaque côté avant la fusion (on garde
    la 1ère occurrence) et on le signale explicitement plutôt que de laisser
    le chiffre affiché mentir en silence)."""
    df_manifeste_cont = df_manifeste_cont.copy()
    df_discharge = df_discharge.copy()
    if "_CONT_norm" not in df_manifeste_cont.columns:
        df_manifeste_cont["_CONT_norm"] = ""
    if "No_Conteneur" not in df_discharge.columns:
        df_discharge["No_Conteneur"] = ""

    set_m = set(df_manifeste_cont["_CONT_norm"]) - {""}
    set_d = set(df_discharge["No_Conteneur"]) - {""}

    manquants_discharge = set_m - set_d   # dans le manifeste, absents du discharging summary
    manquants_manifeste = set_d - set_m   # dans le discharging summary, absents du manifeste
    communs = set_m & set_d

    df_manquants_discharge = df_manifeste_cont[df_manifeste_cont["_CONT_norm"].isin(manquants_discharge)].copy()
    df_manquants_manifeste = df_discharge[df_discharge["No_Conteneur"].isin(manquants_manifeste)].copy()

    df_m_communs = df_manifeste_cont[df_manifeste_cont["_CONT_norm"].isin(communs)]
    df_d_communs = df_discharge[df_discharge["No_Conteneur"].isin(communs)]

    dup_m = df_m_communs["_CONT_norm"][df_m_communs["_CONT_norm"].duplicated()].unique().tolist()
    dup_d = df_d_communs["No_Conteneur"][df_d_communs["No_Conteneur"].duplicated()].unique().tolist()
    dup_warning = None
    if dup_m or dup_d:
        parts = []
        if dup_m:
            parts.append(f"{len(dup_m)} conteneur(s) en double côté manifeste ({', '.join(dup_m[:8])}{'…' if len(dup_m) > 8 else ''})")
        if dup_d:
            parts.append(f"{len(dup_d)} conteneur(s) en double côté Discharging Summary ({', '.join(dup_d[:8])}{'…' if len(dup_d) > 8 else ''})")
        dup_warning = (
            "⚠️ Numéros de conteneur dupliqués détectés — seule la 1ère occurrence de chaque côté "
            "a été rapprochée (les doublons ne sont pas comparés) : " + " ; ".join(parts)
        )
        df_m_communs = df_m_communs.drop_duplicates(subset="_CONT_norm", keep="first")
        df_d_communs = df_d_communs.drop_duplicates(subset="No_Conteneur", keep="first")

    merged = pd.merge(
        df_m_communs, df_d_communs,
        left_on="_CONT_norm", right_on="No_Conteneur", suffixes=("_manifeste", "_discharge"),
    )
    if not merged.empty:
        merged["Poids_manifeste_kg"] = pd.to_numeric(merged["Weight(ton)"], errors="coerce") * 1000
        merged["Poids_discharge_kg"] = pd.to_numeric(merged["Kilogr"], errors="coerce")
        merged["Ecart_kg"] = merged["Poids_manifeste_kg"] - merged["Poids_discharge_kg"]
        merged["Ecart_pct"] = (merged["Ecart_kg"] / merged["Poids_discharge_kg"].replace(0, pd.NA)) * 100
        merged = merged.reindex(merged["Ecart_kg"].abs().sort_values(ascending=False).index)

    return df_manquants_discharge, df_manquants_manifeste, merged, dup_warning


# ---------------------------------------------------------------------------
# Export générique des rapports d'écarts (étapes 2 et 3)
# ---------------------------------------------------------------------------
def build_report_workbook_bytes(sheets: dict, title_lines=None) -> io.BytesIO:
    """sheets : dict {nom_onglet: DataFrame} → classeur Excel stylé (même
    mise en forme que le reste de l'app, via write_sheet). Retire les
    colonnes techniques (préfixées '_') avant écriture."""
    wb = Workbook()
    wb.remove(wb.active)
    for name, df in sheets.items():
        df_clean = df.drop(columns=[c for c in df.columns if str(c).startswith("_")], errors="ignore")
        ws = wb.create_sheet(str(name)[:31])
        data = df_clean.reset_index(drop=True) if df_clean.columns.size else pd.DataFrame({"Info": ["Aucun élément"]})
        write_sheet(ws, data, title_lines=title_lines)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Validation des écarts (étape 2) — un seul fichier corrigé et à jour
# ---------------------------------------------------------------------------
ADDED_FILL = "FDE9D9"    # orange pastel — ligne ajoutée depuis le manifeste (absente de la liste reçue)
FLAGGED_FILL = "F8D7DA"  # rouge pastel — ligne de la liste reçue non retrouvée dans les manifestes traités


def _apply_row_fills(ws, header_row_idx: int, fills: list):
    """Applique une couleur de fond (ou aucune, si None) à chaque ligne de
    données, dans l'ordre d'écriture — fills[i] correspond à la (i+1)-ème
    ligne de données sous l'en-tête."""
    from openpyxl.styles import PatternFill
    for i, color in enumerate(fills):
        if not color:
            continue
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for cell in ws[header_row_idx + 1 + i]:
            cell.fill = fill


def build_liste_corrigee_workbook_bytes(prov: dict, previs: dict, resultats: dict, navire: str, voyage: str) -> io.BytesIO:
    """Produit un classeur unique = la liste provisoire reçue (prov), mise à
    jour directement à partir du rapprochement déjà calculé (resultats,
    sortie de reconcile_bl pour chaque onglet) :
    - B/L des manifestes absents de la liste provisoire → AJOUTÉS en fin
      d'onglet, dans les colonnes de la liste provisoire (alignement par nom
      de colonne — les colonnes du manifeste sans équivalent dans la liste
      provisoire sont omises, pas de colonne inventée), surlignage ORANGE.
    - B/L de la liste provisoire non retrouvés dans les manifestes déjà
      traités → signalés EN PLACE (jamais supprimés : le port correspondant
      n'a peut-être simplement pas encore été traité), surlignage ROUGE.
    Ne couvre que ce rapprochement (étape 2, structure B/L commune entre
    manifeste et liste provisoire) — pas le Discharging Container Summary
    (étape 3), de nature différente (écarts de poids par conteneur, pas une
    liste de B/L à compléter)."""
    title = [
        f"Navire : {navire}", f"Voyage : {voyage}",
        "Liste provisoire mise à jour depuis les manifestes déjà structurés",
        "🟠 orange = ajouté depuis le manifeste (absent de la liste reçue)  |  "
        "🔴 rouge = présent dans la liste reçue mais non retrouvé dans les manifestes déjà traités (à vérifier)",
    ]
    wb = Workbook()
    wb.remove(wb.active)
    for sheet in ("RORO", "CONTENEUR", "BB"):
        df_prov = prov.get(sheet, pd.DataFrame())
        manquants, en_trop, _n_communs = resultats.get(sheet, (pd.DataFrame(), pd.DataFrame(), 0))

        base_cols = [c for c in df_prov.columns if not str(c).startswith("_")]
        if not base_cols:
            # Aucune liste provisoire pour cet onglet (absente du fichier
            # uploadé, ou fichier non fourni) — tout le contenu vient du
            # manifeste, dans les colonnes du gabarit Reporting.
            base_cols = [c for c in previs.get(sheet, pd.DataFrame()).columns if not str(c).startswith("_")]

        base = df_prov[base_cols].copy() if not df_prov.empty else pd.DataFrame(columns=base_cols)
        # en_trop préserve l'index d'origine de df_prov (reconcile_bl filtre
        # par .isin() sans reset_index) — on l'utilise directement pour
        # savoir quelles lignes de base signaler, sans re-matching.
        flagged_idx = set(en_trop.index)
        fills = [FLAGGED_FILL if idx in flagged_idx else None for idx in base.index]

        added = manquants.reindex(columns=base_cols, fill_value="") if not manquants.empty else pd.DataFrame(columns=base_cols)
        fills += [ADDED_FILL] * len(added)

        n_ajoutes = manquants["_BL_norm"].nunique() if not manquants.empty else 0
        n_signales = en_trop["_BL_norm"].nunique() if not en_trop.empty else 0
        title_sheet = title + [
            f"Cet onglet ({sheet}) : {n_ajoutes} B/L ajouté(s) en orange, {n_signales} B/L signalé(s) en rouge."
        ]

        ws = wb.create_sheet(sheet)
        if base.empty and added.empty:
            write_sheet(ws, pd.DataFrame({"Info": ["Aucune donnée (ni liste provisoire, ni manifeste traité)"]}), title_lines=title_sheet)
            continue
        final = pd.concat([base.reset_index(drop=True), added.reset_index(drop=True)], ignore_index=True)
        header_row_idx = write_sheet(ws, final, title_lines=title_sheet)
        _apply_row_fills(ws, header_row_idx, fills)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
