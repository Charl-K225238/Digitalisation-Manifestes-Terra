"""
Prototype - Parseur de manifestes cargo Grimaldi (format PBREPORT)
Approche: parsing deterministe base sur les positions de colonnes (pipes)
plutot que LLM, car le rapport est genere par machine avec un gabarit fixe.
"""
import re
import pdfplumber
import pandas as pd

BL_RE = re.compile(r'^\[?([A-Z]{0,3}\d{5,})\]?(\[T\])?$')
# Forme canonique observée sur tous les B/L Grimaldi réels rencontrés à ce
# jour (1 lettre + 9 chiffres, ex. "S330078443" — RORO/BB ont le même motif
# une fois le tiret retiré par normalize_bl côté reporting_builder.py) :
# sert UNIQUEMENT de garde-fou de signalement (bl_format_anomalies), jamais
# à filtrer/rejeter un B/L extrait — BL_RE reste volontairement plus large
# (0-3 lettres, 5+ chiffres) car resserrer l'extraction risquerait de perdre
# de vrais B/L à un format encore non rencontré (voir audit 17/08 : limite
# connue "LOS42133" — une référence courte à 3 lettres/5 chiffres peut être
# extraite à tort comme B/L juste avant le vrai B/L entre crochets qui suit
# — non corrigé côté extraction faute d'assez d'exemples pour calibrer sans
# risque de faux négatifs, mais désormais détecté et signalé à l'agent).
BL_CANONICAL_RE = re.compile(r'^(?:[A-Z]\d{8,10}|[A-Z]{2,5}\d{4,8})$')
# Deux formes canoniques légitimes observées : le format Grimaldi standard
# (1 lettre + 8-10 chiffres, ex. "S330078443") ET les références courtes de
# certaines agences locales (préfixe port + numéro, ex. "LOS42108" pour
# Lagos) — confirmées légitimes par l'utilisateur (04/09), pas des anomalies
# à signaler.
CONTAINER_RE = re.compile(r'^CN\s*:\s*(\S+)$')
SEAL_RE = re.compile(r'^SN\s*:\s*(\S+)$')
# Accepte entier, 1 ou 2+ décimales, séparateurs de milliers (virgule ou espace),
# et un suffixe d'unité optionnel tronqué par un retour à la ligne PDF (audit
# 17/08, bug #2 : "36,440.000 K" au lieu de "...KGS" — très frequent sur les
# B/L "LM RoRo" de plusieurs manifestes, ex. GTC0526 Anvers ~50 occurrences).
WEIGHT_RE = re.compile(r'^([\d, ]+(?:\.\d+)?)\s*K?G?S?$', re.I)
CBM_RE = re.compile(r'^([\d,]+\.\d{3})\s*CBM$')
LM_RE = re.compile(r'^([\d,]+\.\d{2})\s*LM$')
DATE_RE = re.compile(r'DATED\s+([\d/\-]+)')
ORIG_BL_RE = re.compile(r'ORIGINAL BILL OF LADING\s+(\S+)')
FREIGHT_RE = re.compile(r'Freight payable at\s*:\s*(.+)')
HS_CODE_RE = re.compile(r'H\.?S\.?\s*CODE\s*:\s*(\S+)', re.I)
MODEL_YEAR_RE = re.compile(r'Model\s*Year\s*:?\s*(\d{4})|MODEL\s*:\s*(\d{4})', re.I)
TRANSIT_TO_RE = re.compile(r"TRANSIT TO\s*:?\s*([A-Z][A-Za-z .,'\-]{1,40})", re.I)
LOCAL_AREA_RE = re.compile(
    r'ABIDJAN|IVORY COAST|COTE D|CÔTE D|C\u2019?OTE D|TREICHVILLE|COCODY|YOPOUGON|MARCORY|'
    r'PLATEAU|ADJAME|KOUMASSI|PORT[- ]?BOUET|BINGERVILLE|ANYAMA|RIVIERA|ANGRE|ATTECOUBE|ABOBO',
    re.I)
FT_SIZE_RE = re.compile(r'(\d{2})\s*ft', re.I)
# --- "Bebe au dos" (02/09 v7) : engin(s) porte(s) par l'unite principale du B/L,
# identifie(s) par leur propre numero de chassis (matricule), a faire remonter
# en ligne(s) supplementaire(s) a la suite du B/L principal.
TRAILER_ATTACHED_RE = re.compile(r'^\+\s*(\d+)\s+TRAILER\s+ATTACHED', re.I)
STACKED_WITH_RE = re.compile(r'S\.?T\.?B\.?\s*STACKED\s+WITH', re.I)
STACKED_UNIT_RE = re.compile(r'^(\d+)\s*Unit', re.I)
WITH_CH_RE = re.compile(r'With\s+CH#\s*:?\s*([A-Z0-9]+)', re.I)
MODEL_SREM_RE = re.compile(r'MODEL\s+SREM\s*:\s*(.+)', re.I)
# --- Robustesse (audit 17/08, jamais reintegre au fichier deploye - reapplique 02/09) ---
SAID_TO_CONTAIN_RE = re.compile(r'SAID\s+TO\s+CONTAIN', re.I)
WEIGHT_VOLUME_RE = re.compile(r'([\d,]+\.\d+)\s*KGS?\s*-\s*([\d,]+\.\d+)\s*M3', re.I)
TOTAL_WEIGHT_RE = re.compile(r'(?:GROSS|TOTAL)\s+WEIGHT\s*[:=]\s*([\d,]+\.?\d*)\s*(KGS?|MT)?', re.I)
COLOR_RE = re.compile(r'COLOR\s*:\s*([A-Za-z /]+?)(?:\s{2,}|$|\||H\.?S)', re.I)
ENGINE_NO_RE = re.compile(r'Engine\s*No\.?\s*[:.]?\s*([A-Z0-9]{5,})', re.I)
VEHICLE_TYPE_RE = re.compile(r"Van\(s\)|Car\(s\)|RoRo|Tractor", re.I)
# Marques automobiles reconnues pour extraction automatique depuis la description
MARQUE_RE = re.compile(
    r'\b(TOYOTA|KIA|NISSAN|FORD|HYUNDAI|MITSUBISHI|MAZDA|HONDA|CHEVROLET|'
    r'RENAULT|PEUGEOT|CITROEN|BMW|MERCEDES(?:[- ]BENZ)?|VOLKSWAGEN|VW|'
    r'VOLVO|SCANIA|ISUZU|SUZUKI|DAIHATSU|SUBARU|JEEP|LAND ROVER|RANGE ROVER|'
    r'LEXUS|INFINITI|DACIA|OPEL|FIAT|IVECO|MAN\b|DAF|HINO|TATA|MAHINDRA|'
    r'GEELY|BYD|JAC|CHERY|MG|SSANGYONG|FOTON|JMC|SINOTRUK|YUTONG)\b',
    re.I
)
# Mots qui suivent une marque mais ne sont PAS un modèle (à ignorer)
_NON_MODEL = frozenset({
    "NEW", "USED", "CAR", "CARS", "VAN", "VANS", "VEHICLE", "VEHICLES",
    "TRUCK", "TRUCKS", "BUS", "BUSES", "RORO", "CARGO", "HEAVY",
})
FOOTER_MARKERS = ("Totals For", "Grand Totals", "Summary Totals", "End Of Report")
# En-têtes de tableau répétés en haut de chaque page -> à ignorer entièrement.
# BUG CRITIQUE trouvé et corrigé le 02/09 : "Grimaldi Deep Sea S.p.A." a été
# retiré de cette liste. C'est a la fois le texte du bandeau de titre de page
# ET la 2e ligne standard de l'adresse du shipper sur PRESQUE CHAQUE B/L
# ("as agent of / Grimaldi Deep Sea S.p.A. / Piazza..."). Le test `any(marker
# in cols for marker in HEADER_ROW_MARKERS)` supprimait donc ENTIEREMENT
# toute ligne de donnees ou cette adresse partageait la ligne PDF avec une
# autre colonne utile (description, CN:/SN:, quantite...) — verifie : 593
# lignes avec donnee reelle perdue sur un seul manifeste de 114 pages
# (GRA0526 Dakar). "CARGO MANIFEST" a lui seul identifie deja de maniere
# fiable le bandeau de titre (jamais present dans une donnee de B/L) et
# suffit a filtrer ces lignes sans ce risque.
HEADER_ROW_MARKERS = (
    "B/L No.", "SHIPPER(SH), CONSIGNEE(CN), NOTIFY(NO", "Marks And Nos.;",
    "Numbers And Kind Of Packages;", "Name Of Ship And Voyage No.",
    "Nationality Of Ship", "Name Of Master", "Place Of Receipt",
    "CARGO MANIFEST", "Move Type", "Origin Port",
    "Port Where & When Report is made", "Weight(Kgs)", "Charge Information",
)


def extract_rows(pdf_path, progress_cb=None):
    """Extrait toutes les lignes de toutes les pages, splittees par colonnes (pipe).

    progress_cb(page_courante, total_pages), si fourni, est appele apres
    chaque page — permet a l'appelant (app Streamlit) d'afficher une
    progression reelle pendant le traitement d'un manifeste long.

    Perf (02/09) : `page.flush_cache()` est appele apres chaque page. Sans
    cela, pdfplumber conserve en memoire les objets de toutes les pages deja
    lues (chars/lignes), et le temps par page croit avec le nombre de pages
    deja traitees — mesure sur un manifeste reel de 114 pages : 112s sans
    flush_cache() contre 30s avec, pour un texte extrait strictement
    identique (verifie octet pour octet). Aucun changement de comportement
    d'extraction, uniquement la gestion memoire de pdfplumber."""
    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        for pno, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            for line in text.split("\n"):
                cols = [c.strip() for c in line.split("|")]
                # retire la 1ere colonne vide (avant le premier pipe) et la derniere si vide
                if cols and cols[0] == "":
                    cols = cols[1:]
                if cols and cols[-1] == "":
                    cols = cols[:-1]
                rows.append((pno, cols))
            page.flush_cache()
            if progress_cb is not None:
                progress_cb(pno, total)
    return rows


def is_separator(cols):
    if not cols:
        return True
    joined = "".join(cols).strip()
    return joined == "" or set(joined) <= {"-"}


def is_footer(cols):
    joined = " ".join(cols)
    return any(m in joined for m in FOOTER_MARKERS)


def get(cols, i):
    return cols[i].strip() if i < len(cols) else ""


def parse_manifest(pdf_path, source_label, progress_cb=None):
    rows = extract_rows(pdf_path, progress_cb=progress_cb)

    context = {
        "vessel_voyage": "", "move_type": "", "origin_port": "",
        "port_of_loading": "", "port_of_discharge": "",
    }

    records = []
    current = None  # dict pour le B/L en cours
    state = None  # "SH" | "CN" | "NO"

    def flush():
        nonlocal current
        if current:
            records.append(current)
        current = None

    for idx, (pno, cols) in enumerate(rows):
        if not cols or is_separator(cols):
            continue
        if is_footer(cols):
            continue
        if any(marker in cols for marker in HEADER_ROW_MARKERS):
            continue

        # --- Detection / mise a jour du contexte de section ---
        # Ligne "Nom navire / Lieu rapport / Sens / Port origine / Date depart POL".
        # Reperee par structure (colonne "Sens" = "H : H" / "P : P" + colonne 0
        # au format "NAVIRE : VOYAGE"), PAS par le nom d'un navire particulier
        # (un manifeste peut concerner n'importe quel navire de la flotte).
        if (get(cols, 0) and ":" in get(cols, 0) and len(cols) >= 3
                and get(cols, 2) in ("H : H", "P : P")):
            context["vessel_voyage"] = get(cols, 0)
            context["move_type"] = get(cols, 2)
            context["origin_port"] = get(cols, 3) if len(cols) > 3 else ""
            continue
        if get(cols, 0) == "Italy" and len(cols) >= 4:
            context["port_of_loading"] = get(cols, 3)
            context["port_of_discharge"] = get(cols, 4) if len(cols) > 4 else ""
            continue

        col0 = get(cols, 0)
        m = BL_RE.match(col0)
        # Fallback : certaines pages n'ont pas de séparateurs pipe —
        # le numéro de B/L et le contenu de la ligne se retrouvent tous
        # dans cols[0]. On tente d'extraire un numéro de B/L en début
        # de col0 si le match direct a échoué et que la ligne est courte.
        if not m and col0 and len(cols) == 1:
            fb = re.match(r'^(\[?[A-Z]{0,3}\d{6,}\]?(?:\[T\])?)(?:\s|$)', col0)
            if fb:
                m = BL_RE.match(fb.group(1))
        if m:
            # nouveau B/L -> on cloture le precedent
            flush()
            current = {
                "source_file": source_label,
                "vessel_voyage": context["vessel_voyage"],
                "move_type": context["move_type"],
                "origin_port": context["origin_port"],
                "port_of_loading": context["port_of_loading"],
                "port_of_discharge": context["port_of_discharge"],
                "bl_number": m.group(1),
                "transshipment": bool(m.group(2)),
                "shipper_name": "", "shipper": [], "consignee_name": [], "consignee_address": [],
                "notify_name": [], "notify_address": [],
                "items": [],  # liste d'items {qty, type_raw, weight, tare, cbm, lm, chassis, container_no, seal_no}
                "freight_payable_at": "", "original_bl_ref": "", "original_bl_date": "",
                "raw_desc_lines": [],
                "_pending_piggyback": None, "_awaiting_stacked_qty": False,
                "_collecting_piggyback_desc": False, "_suppress_qty": False,
            }
            state = "SH"

        if current is None:
            continue  # lignes hors-B/L (avant le 1er record) -> ignorees

        c1 = get(cols, 1)  # colonne SHIPPER/CONSIGNEE/NOTIFY
        c2 = get(cols, 2)  # colonne Unit No.
        c3 = get(cols, 3)  # colonne Description
        c4 = get(cols, 4)  # colonne Weight
        c5 = get(cols, 5)  # colonne Measurement

        def active_item():
            if not current["items"]:
                current["items"].append({"qty": 1, "type_raw": "", "weight": None, "tare": None,
                                          "cbm": None, "lm": None, "chassis": [],
                                          "container_no": [], "seal_no": [], "_is_container_slot": False})
            return current["items"][-1]

        def container_target():
            """Item qui doit recevoir le prochain CN:/SN: rencontré.

            Cas piégeux (14/08 v6) : le manifeste liste souvent D'ABORD toutes
            les descriptions de N conteneurs identiques ("1-40 ft. High Cube"
            x5), PUIS, plus bas, le bloc de leurs CN:/SN:/poids un par un.
            En prenant toujours le "dernier item créé" (comportement
            précédent), les N conteneurs finissaient tous rattachés au même
            item (souvent une ligne de description du contenu créée entre
            les deux blocs, ex. "599 PIECES Pallet EXPORT...") : poids
            écrasé au lieu d'être réparti, conteneurs listés en doublon sur
            une seule ligne. On cible ici le premier "emplacement conteneur"
            (item créé depuis une description avec taille en pieds) encore
            sans numéro de conteneur, dans l'ordre d'apparition — ce qui
            réplique la correspondance 1 pour 1 réelle du manifeste."""
            for it in current["items"]:
                if it.get("_is_container_slot") and not it["container_no"]:
                    return it
            return active_item()

        # --- colonne 1 : shipper / consignee / notify ---
        if c1:
            if c1.startswith("SH:"):
                state = "SH"
                name = c1[3:].strip()
                current["shipper_name"] = name
                current["shipper"].append(name)
            elif c1.startswith("CN:"):
                state = "CN"
                current["consignee_name"].append(c1[3:].strip())
            elif re.match(r'^NO\[\d+\]:', c1):
                state = "NO"
                current["notify_name"].append(re.sub(r'^NO\[\d+\]:', '', c1).strip())
            else:
                if state == "SH":
                    current["shipper"].append(c1)
                elif state == "CN":
                    current["consignee_address"].append(c1)
                elif state == "NO":
                    current["notify_address"].append(c1)

        # --- colonne 2 : conteneur / scellé / chassis ---
        # Conteneur/scellé : rattachés à l'"emplacement conteneur" en cours de
        # remplissage (container_target, voir plus haut), pas systématiquement
        # au dernier item créé. Châssis (véhicules) : comportement inchangé,
        # un seul item par B/L dans ce cas donc pas d'ambiguïté.
        if c2:
            mc = CONTAINER_RE.match(c2)
            ms = SEAL_RE.match(c2)
            if mc:
                tgt = container_target()
                tgt["container_no"].append(mc.group(1))
                current["_last_touched"] = tgt
            elif ms:
                tgt = current.get("_last_touched") or active_item()
                tgt["seal_no"].append(ms.group(1))
                current["_last_touched"] = tgt
            elif c2 == "CHASSIS NOS :":
                pass
            elif re.match(r'^[A-Z0-9]{10,}$', c2):
                active_item()["chassis"].append(c2)

        # --- colonne 3 : description / type de colis (créé un nouvel item) / infos service ---
        if c3:
            # --- "Bebe au dos" : remorque attelée ("+1 TRAILER ATTACHED" /
            # "With CH# :...") ou engin empilé ("S.T.B. STACKED WITH :" / "N
            # Unit" / description / "With CH# :..."). Chacun devient un item
            # V a part entiere, avec son propre chassis, a la suite du B/L
            # principal (voir records_to_dataframe : type_code force a "V").
            tam = TRAILER_ATTACHED_RE.match(c3)
            wcm = WITH_CH_RE.search(c3)
            swm = STACKED_WITH_RE.search(c3)
            sum_m = STACKED_UNIT_RE.match(c3) if current.get("_awaiting_stacked_qty") else None
            msm = MODEL_SREM_RE.match(c3) if current.get("_pending_piggyback") else None
            fm = FREIGHT_RE.search(c3)
            om = ORIG_BL_RE.search(c3)
            dm = DATE_RE.search(c3)
            # (audit 17/08, bug #5 + variante "SAID TO"/"CONTAIN" coupes sur 2
            # lignes par le retour a la ligne PDF, ex. GTC0526 Anvers "49 X 20'
            # CONTAINERS SAID TO" / "CONTAIN") : toute etiquette qui se termine
            # par ":" sans valeur sur la meme ligne ("TOTAL NUMBER OF
            # CONTAINERS:") ou par "SAID TO" annonce que la ligne suivante est
            # une VALEUR de continuation, pas un nouveau colis/vehicule.
            stcm = (SAID_TO_CONTAIN_RE.search(c3) or c3.rstrip().endswith(':')
                    or c3.strip().upper().endswith('SAID TO'))
            # NB (14/08 v6) : "Container"/"Tank"/"ft\." ajoutés après avoir
            # constaté que des conteneurs vides ("1-20 ft. Tank Container",
            # ex. GTC0526 Amsterdam, ~28 unités sous un seul B/L) ne
            # matchaient aucun des mots-clés précédents -> aucun item créé
            # -> tous les CN:/SN: suivants s'accrochaient au même item par
            # défaut (fusion silencieuse de 28 conteneurs en 1 seule ligne,
            # poids écrasé au lieu d'être sommé). Un conteneur vide reste une
            # ligne valide à part entière, pas une anomalie à filtrer.
            # NB (02/09 v7, audit 17/08 reappliqué) : "\bCar(?:\(s\))?\b" au
            # lieu de "Car" nu — "Car" nu matchait n'importe quel mot contenant
            # la sous-chaine ("CARTONS", "CARRIER"...), creant des items
            # fantomes sur de simples lignes de contenu.
            qty_m = re.match(
                r'^(\d+)[\s\-]+(.*(?:Van|Cargo|Cube|\bCar(?:\(s\))?\b|LM RoRo|PIECE|CRATE|'
                r'Tractor|PACKAGE|Container|Tank|ft\.).*)$', c3, re.I)

            if tam:
                pb = {"qty": int(tam.group(1)) or 1, "type_raw": "Remorque attelée",
                      "weight": None, "tare": None, "cbm": None, "lm": None,
                      "chassis": [], "container_no": [], "seal_no": [],
                      "_is_container_slot": False, "_piggyback": True,
                      "_piggyback_kind": "Attelée (remorque)"}
                current["items"].append(pb)
                current["_pending_piggyback"] = pb
                current["_last_touched"] = None
            elif wcm and current.get("_pending_piggyback"):
                current["_pending_piggyback"]["chassis"].append(wcm.group(1))
                current["_collecting_piggyback_desc"] = False
            elif swm:
                current["_awaiting_stacked_qty"] = True
            elif sum_m:
                pb = {"qty": int(sum_m.group(1)) or 1, "type_raw": "",
                      "weight": None, "tare": None, "cbm": None, "lm": None,
                      "chassis": [], "container_no": [], "seal_no": [],
                      "_is_container_slot": False, "_piggyback": True,
                      "_piggyback_kind": "Empilée (bébé au dos)"}
                current["items"].append(pb)
                current["_pending_piggyback"] = pb
                current["_awaiting_stacked_qty"] = False
                current["_collecting_piggyback_desc"] = True
            elif msm:
                pending = current["_pending_piggyback"]
                pending["type_raw"] = (pending["type_raw"] + " " + msm.group(1).strip()).strip()
            elif current.get("_collecting_piggyback_desc"):
                pending = current["_pending_piggyback"]
                pending["type_raw"] = (pending["type_raw"] + " " + c3).strip()
            elif stcm:
                # Marqueur "CONTAINER(S) SAID TO CONTAIN" : tout ce qui suit
                # decrit le CONTENU du conteneur deja cree juste avant, pas
                # un nouveau colis/vehicule — sauf une nouvelle taille en
                # pieds ("N-NN ft. ...") qui signale un conteneur suivant
                # legitime dans le meme B/L (voir plus bas, qty_m).
                current["_suppress_qty"] = True
            elif fm:
                current["freight_payable_at"] = fm.group(1).strip()
                current["_suppress_qty"] = False
            elif om:
                current["original_bl_ref"] = om.group(1)
                current["_suppress_qty"] = False
            elif dm:
                current["original_bl_date"] = dm.group(1)
            elif qty_m:
                type_raw = qty_m.group(2).strip()
                is_container_decl = bool(FT_SIZE_RE.search(type_raw))
                if current.get("_suppress_qty") and not is_container_decl:
                    # (audit 17/08, bug #3) Ligne de CONTENU d'un conteneur
                    # deja cree ("17 PACKAGES", "200 Cartons de ...") — pas
                    # un nouvel item. Gardee en texte descriptif seulement.
                    current["raw_desc_lines"].append(c3)
                else:
                    current["_suppress_qty"] = False
                    current["items"].append({
                        "qty": int(qty_m.group(1)), "type_raw": type_raw,
                        "weight": None, "tare": None, "cbm": None, "lm": None,
                        "chassis": [], "container_no": [], "seal_no": [],
                        # Emplacement conteneur (taille en pieds explicite) vs.
                        # simple ligne de description de contenu (ex. "PIECES
                        # Pallet EXPORT...") — voir container_target() plus haut.
                        "_is_container_slot": is_container_decl,
                    })
                    current["_last_touched"] = None
            elif c3 == "TARE":
                pass  # le tare est en colonne weight, gere plus bas
            elif c3 == "Service B/L":
                pass  # marqueur de type de B/L, pas un type de colis
            else:
                # (audit 17/08, bugs #1/#6) Poids (+volume) parfois donnés en
                # texte libre dans la description plutot que dans la colonne
                # Weight : "4699.57 KG - 9.616 M3", "GROSS WEIGHT:76.03MT",
                # "TOTAL WEIGHT=34,870KGS". Complete l'item seulement si son
                # poids/volume n'est pas deja connu (n'ecrase jamais une
                # valeur deja lue depuis la colonne Weight/Measurement).
                wvm = WEIGHT_VOLUME_RE.search(c3)
                twm = TOTAL_WEIGHT_RE.search(c3)
                target = current.get("_last_touched") or (current["items"][-1] if current["items"] else None)
                if wvm and target is not None:
                    if target["weight"] is None:
                        target["weight"] = float(wvm.group(1).replace(",", ""))
                    if target["cbm"] is None:
                        target["cbm"] = float(wvm.group(2).replace(",", ""))
                elif twm and target is not None and target["weight"] is None and twm.group(2):
                    # Unite exigee explicitement sur la meme ligne : quand
                    # l'unite est elle-meme coupee sur la ligne PDF suivante
                    # ("TOTAL GROSS WEIGHT: 1326.969" / "METRIC TONS"), on ne
                    # devine PAS KG par defaut — un poids en tonnes pris pour
                    # des kg serait une erreur x1000 silencieuse, pire que de
                    # laisser le poids vide pour completion par l'agent.
                    val = float(twm.group(1).replace(",", ""))
                    unit = twm.group(2).upper()
                    target["weight"] = val * 1000 if unit == "MT" else val
                current["raw_desc_lines"].append(c3)

        # --- colonne 4 : poids (gross ou tare selon description) ---
        # Rattaché au même item que le dernier CN:/SN: rencontré s'il y en a
        # un pour ce B/L (cas conteneurs multiples groupés), sinon à l'item
        # actif comme avant (cas simple, 1 seul item par B/L).
        if c4:
            # Normalisation : supprime séparateurs de milliers (virgule ou espace)
            # avant d'appliquer la regex, pour accepter "15,000.00" et "15 000.00".
            c4_norm = c4.replace(",", "").replace(" ", "")
            wm = WEIGHT_RE.match(c4_norm)
            if wm:
                val = float(wm.group(1).replace(",", "").replace(" ", ""))
                # Si c3 vient de créer un nouvel item (qty_m), le poids doit
                # rester rattaché à ce nouvel item (comportement voulu quand
                # description + poids sont sur la même ligne). Si _last_touched
                # est None (reset par qty_m) → active_item() est le nouvel item.
                target = current.get("_last_touched") or active_item()
                if c3 == "TARE":
                    target["tare"] = val
                else:
                    target["weight"] = val

        # --- colonne 5 : CBM ou LM (même logique de rattachement que le poids) ---
        if c5:
            c5_norm = c5.replace(" ", "")  # retire les espaces (séparateurs de milliers)
            cm = CBM_RE.match(c5_norm)
            lmm = LM_RE.match(c5_norm)
            if cm or lmm:
                target = current.get("_last_touched") or active_item()
                if cm:
                    target["cbm"] = float(cm.group(1).replace(",", ""))
                else:
                    target["lm"] = float(lmm.group(1).replace(",", ""))

    flush()
    return records


def simplify_type_colis(type_raw):
    """Réduit le type de colis à l'information essentielle.
    Conteneur -> taille en ft ('20'/'40'). Véhicule -> catégorie courte.
    Divers -> 'Colis'/'Caisse'. Retourne (libelle, code C/V/D)."""
    ft = FT_SIZE_RE.search(type_raw)
    if ft:
        return ft.group(1), "C"
    if re.search(r'Small Van', type_raw, re.I):
        return "Van léger", "V"
    if re.search(r'Big Van', type_raw, re.I):
        return "Van lourd", "V"
    if re.search(r'Car\(s\)', type_raw, re.I):
        return "Voiture", "V"
    if re.search(r'RoRo', type_raw, re.I):
        return "RoRo", "V"
    if re.search(r'Tractor|Construction Equip', type_raw, re.I):
        return "Engin lourd", "V"
    if re.search(r'CRATE|PIECE|PACKAGE', type_raw, re.I):
        return "Colis", "D"
    return (type_raw[:20] if type_raw else ""), "D"


COMMUNES_ABIDJAN = [
    "TREICHVILLE", "COCODY", "YOPOUGON", "MARCORY", "PLATEAU", "ADJAME",
    "KOUMASSI", "PORT-BOUET", "PORT BOUET", "BINGERVILLE", "ANYAMA",
    "ATTECOUBE", "ABOBO", "RIVIERA", "ANGRE", "AKOUEDO", "SONGON", "TREICHIVLLE",
]
# Autres villes ivoiriennes frequemment vues dans les adresses destinataire
# (hors Abidjan) — pattern avant nom canonique, ordre : cas les plus
# specifiques d'abord (ex. Grand-Bassam avant un eventuel "\bMAN\b" trop large).
VILLES_CI_PATTERNS = [
    (re.compile(r"GRAND[- ]?BASSAM", re.I), "Grand-Bassam"),
    (re.compile(r"YAMOUSSOUKRO", re.I), "Yamoussoukro"),
    (re.compile(r"SAN[- ]?PEDRO", re.I), "San-Pédro"),
    (re.compile(r"BOUAKE", re.I), "Bouaké"),
    (re.compile(r"KORHOGO", re.I), "Korhogo"),
    (re.compile(r"DALOA", re.I), "Daloa"),
    (re.compile(r"GAGNOA", re.I), "Gagnoa"),
    (re.compile(r"ABENGOUROU", re.I), "Abengourou"),
    (re.compile(r"DABOU", re.I), "Dabou"),
    (re.compile(r"\bMAN\b", re.I), "Man"),
]
# Pays reconnus dans les adresses/mentions de transit. Les variantes
# orthographiques/typos vues sur de vrais manifestes (ex. "BURKINO" au lieu de
# "BURKINA", "FASSO" au lieu de "FASO") sont couvertes par des patterns larges
# (racine du mot) plutot qu'une liste figee de graphies exactes — a completer
# si de nouveaux pays/typos apparaissent dans de futurs manifestes.
COUNTRY_PATTERNS = [
    (re.compile(r"IVORY COAST|COTE D.?IVOIRE|C.TE D.IVOIRE", re.I), "Côte d'Ivoire"),
    (re.compile(r"\bMALI\b", re.I), "Mali"),
    # "BURKIN" seul couvre Burkina/Burkino/Burkinabe (typos observees dans les PDF sources)
    (re.compile(r"BURKIN\w*", re.I), "Burkina Faso"),
    (re.compile(r"\bNIGERIA\b", re.I), "Nigéria"),
    (re.compile(r"\bNIGER\b", re.I), "Niger"),
    (re.compile(r"SENEGAL", re.I), "Sénégal"),
    (re.compile(r"\bGHANA\b", re.I), "Ghana"),
    (re.compile(r"\bTOGO\b", re.I), "Togo"),
    (re.compile(r"\bBENIN\b", re.I), "Bénin"),
    (re.compile(r"GUINEA[- ]?BISSAU|GUIN.E[- ]?BISSAU", re.I), "Guinée-Bissau"),
    (re.compile(r"\bGUINEA\b|\bGUIN.E\b", re.I), "Guinée"),
    (re.compile(r"\bLIBERIA\b", re.I), "Libéria"),
    (re.compile(r"SIERRA LEONE", re.I), "Sierra Leone"),
    (re.compile(r"\bCHAD\b|\bTCHAD\b", re.I), "Tchad"),
    (re.compile(r"CAMEROON|CAMEROUN", re.I), "Cameroun"),
    (re.compile(r"\bCONGO\b", re.I), "Congo"),
    (re.compile(r"\bGABON\b", re.I), "Gabon"),
    (re.compile(r"MAURITANIA|MAURITANIE", re.I), "Mauritanie"),
    (re.compile(r"MOROCCO|\bMAROC\b", re.I), "Maroc"),
    (re.compile(r"\bALGERIA\b|\bALG.RIE\b", re.I), "Algérie"),
    (re.compile(r"\bTUNISIA\b|\bTUNISIE\b", re.I), "Tunisie"),
    (re.compile(r"U\.?S\.?A\.?\b|UNITED STATES", re.I), "États-Unis"),
    (re.compile(r"\bITALY\b", re.I), "Italie"),
    (re.compile(r"PORTUGAL", re.I), "Portugal"),
    (re.compile(r"\bSPAIN\b|ESPAGNE", re.I), "Espagne"),
    (re.compile(r"\bFRANCE\b", re.I), "France"),
    (re.compile(r"GERMANY|ALLEMAGNE", re.I), "Allemagne"),
    (re.compile(r"BELGIUM|BELGIQUE", re.I), "Belgique"),
    (re.compile(r"NETHERLANDS|PAYS[- ]BAS", re.I), "Pays-Bas"),
    (re.compile(r"UNITED KINGDOM|\bUK\b", re.I), "Royaume-Uni"),
    (re.compile(r"SAUDI ARAB", re.I), "Arabie Saoudite"),
    (re.compile(r"\bUAE\b|EMIRATES|EMIRATS", re.I), "Émirats Arabes Unis"),
    (re.compile(r"SOUTH AFRICA", re.I), "Afrique du Sud"),
    (re.compile(r"\bBRAZIL\b|BR.SIL", re.I), "Brésil"),
    (re.compile(r"\bCHINA\b|\bCHINE\b", re.I), "Chine"),
    (re.compile(r"\bINDIA\b", re.I), "Inde"),
    (re.compile(r"\bTURKEY\b|TURQUIE", re.I), "Turquie"),
    (re.compile(r"LEBANON|LIBAN", re.I), "Liban"),
]


def _normalize_country(text):
    """Cherche un pays reconnu (COUNTRY_PATTERNS) dans un texte libre.
    Retourne le nom canonique en francais, ou "" si aucun pays connu."""
    for pat, name in COUNTRY_PATTERNS:
        if pat.search(text):
            return name
    return ""


def detect_transit(full_desc, consignee_addr, notify_addr):
    """Detecte si la marchandise transite au-dela d'Abidjan (regime transit
    vers pays enclave : Mali, Burkina Faso, Niger...), PAS le transbordement
    navire-a-navire (deja capture par Port_Origine_Transbordement).
    Retourne (bool_transit, pays_detecte_ou_vide, niveau_confiance).

    Le pays capture apres "TRANSIT TO" est normalise via COUNTRY_PATTERNS
    (corrige typos/variantes ex. "BURKINO"->"Burkina Faso", et evite d'afficher
    un mot brut/tronque comme "Burkina" au lieu de "Burkina Faso"). Si le
    texte capture n'est pas un pays connu (ex. "TRANSIT TO SANBRADO", un site
    minier), on cherche un pays connu dans le reste du B/L avant d'abandonner."""
    m = TRANSIT_TO_RE.search(full_desc)
    if m:
        captured = m.group(1).strip()
        pays = _normalize_country(captured) or _normalize_country(full_desc)
        if pays:
            return True, pays, "haute"
        # Mention de transit explicite mais destination non reconnue dans notre
        # liste : on garde le texte brut (mieux que rien) avec confiance
        # abaissee pour signaler a l'agent de verifier/completer.
        return True, captured.title(), "moyenne"
    addr = " ".join(consignee_addr + notify_addr)
    if addr.strip() and not LOCAL_AREA_RE.search(addr):
        # aucune reference locale (Abidjan/quartier/Cote d'Ivoire) dans l'adresse
        return True, "", "faible"
    return False, "", "haute"


def simplify_address(addr):
    """Reduit une adresse brute a l'essentiel : Commune/Ville, Pays.
    Best-effort (liste de communes/villes/pays connus) ; a completer si de
    nouvelles zones apparaissent dans de futurs manifestes."""
    if not addr:
        return ""
    commune = next((c for c in COMMUNES_ABIDJAN if re.search(re.escape(c), addr, re.I)), None)
    if commune or re.search(r'\bABIDJAN\b', addr, re.I):
        ville = "Abidjan"
    else:
        ville = next((name for pat, name in VILLES_CI_PATTERNS if pat.search(addr)), "")
    pays = _normalize_country(addr)
    parts = [p for p in [commune.title() if commune else None, ville, pays] if p]
    return ", ".join(parts)


def extract_marque_modele(text: str):
    """Extrait marque et modèle d'un véhicule depuis le texte descriptif.
    Cherche d'abord dans type_raw (description de l'item), puis dans
    le contexte B/L (raw_desc_lines). Retourne ("", "") si aucune marque
    connue n'est trouvée — évite de polluer les lignes conteneurs/colis."""
    m = MARQUE_RE.search(text)
    if not m:
        return "", ""
    marque = m.group(1).upper().replace("MERCEDES-BENZ", "MERCEDES")
    # Premier token capitalisé après la marque qui n'est pas un mot générique
    after = text[m.end():].strip()
    modele = ""
    for token in after.split():
        # rstrip("(S)") supprimerait chaque caractère de l'ensemble {(,S,)}
        # plutôt que la sous-chaîne "(S)" — on utilise removesuffix pour être précis.
        tok_upper = token.upper().removesuffix("(S)").removesuffix("S")
        if tok_upper and tok_upper[0].isalpha() and tok_upper not in _NON_MODEL:
            modele = token.title()
            break
    return marque.title(), modele


def item_status(type_raw, desc_context):
    """Neuf si 'NEW' mentionné pour cet item (dans son libellé ou le contexte
    descriptif proche), Occasion si 'USED', vide sinon."""
    text = type_raw + " " + desc_context
    if re.search(r'\bNEW\b', text, re.I):
        return "Neuf"
    if re.search(r'\bUSED\b', text, re.I):
        return "Usager"
    return ""


def bl_format_anomalies(bl_numeros) -> list:
    """Retourne, triée, la liste des numéros de B/L extraits qui NE
    correspondent PAS à la forme canonique Grimaldi (BL_CANONICAL_RE) —
    signal pour l'agent qu'une référence a peut-être été confondue avec un
    vrai B/L (voir commentaire BL_CANONICAL_RE), à vérifier sur le PDF
    source avant de faire confiance aux rapprochements Reporting qui
    s'appuient dessus. N'altère jamais l'extraction elle-même : purement
    diagnostique. Accepte n'importe quel itérable de chaînes (typiquement
    df["BL_Numero"].dropna().unique()) ; ignore les valeurs vides."""
    out = []
    for bl in bl_numeros:
        if bl is None or (isinstance(bl, float) and bl != bl):  # None / NaN — pas une anomalie de format
            continue
        s = str(bl).strip()
        if s and not BL_CANONICAL_RE.match(s):
            out.append(s)
    return sorted(set(out))


def records_to_dataframe(records):
    """Vue groupee : 1 ligne = 1 (B/L, Type de colis simplifie).
    Les items de meme type au sein d'un meme B/L sont additionnes (quantite,
    poids, volume)."""
    rows = []
    for r in records:
        full_desc = " | ".join(dict.fromkeys(r["raw_desc_lines"]))
        is_transit, transit_pays, transit_conf = detect_transit(
            full_desc, r["consignee_address"], r["notify_address"])
        nav_m = re.match(r'^(.*?)\s*:\s*(\S+)$', r["vessel_voyage"])
        if nav_m:
            navire = nav_m.group(1).strip()
            voyage = nav_m.group(2).strip()
        else:
            navire = r["vessel_voyage"].strip() or "(navire non détecté)"
            voyage = ""
        addr_simple = simplify_address(", ".join(r["consignee_address"]))

        # Champs au niveau B/L (mêmes pour tous les items de ce B/L)
        nature_bl = "Transb." if r.get("transshipment") else "Import"
        port_dech = r.get("port_of_discharge", "")

        # Année de fabrication : extraite de la description complète du B/L
        year_search = full_desc + " " + " ".join(it["type_raw"] for it in r["items"])
        ym = MODEL_YEAR_RE.search(year_search)
        annee_fab = (ym.group(1) or ym.group(2)) if ym else ""

        # Couleur, Code HS, N° Moteur — extraits de la description globale du B/L
        color_m = COLOR_RE.search(full_desc)
        couleur_bl = color_m.group(1).strip().title() if color_m else ""
        hs_m = HS_CODE_RE.search(full_desc)
        code_hs_bl = hs_m.group(1).strip() if hs_m else ""
        engine_m = ENGINE_NO_RE.search(full_desc)
        no_moteur_bl = engine_m.group(1).strip() if engine_m else ""

        for it in r["items"]:
            statut = item_status(it["type_raw"], full_desc)
            if it.get("_piggyback"):
                # "Bebe au dos" : remorque attelee ou engin empile sur
                # l'unite principale du B/L, identifie par son propre
                # chassis — toujours une ligne Vehicule a part entiere.
                type_simple = it.get("_piggyback_kind") or "Bébé au dos"
                type_code = "V"
                marque, modele = extract_marque_modele(it["type_raw"])
                bebe_au_dos = "Oui"
            else:
                type_simple, type_code = simplify_type_colis(it["type_raw"])
                # Marque/Modèle : cherchée d'abord dans le libellé de l'item,
                # puis dans la description globale du B/L en secours.
                marque, modele = extract_marque_modele(it["type_raw"])
                if not marque:
                    marque, modele = extract_marque_modele(full_desc)
                bebe_au_dos = ""
            rows.append({
                "Fichier": r["source_file"],
                "Navire": navire,
                "Voyage": voyage,
                "Port_Chargement": r.get("port_of_loading", ""),
                "Port_Dechargement": port_dech,
                "BL_Numero": r["bl_number"],
                "Nature_BL": nature_bl,
                "Chargeur_Nom": r["shipper_name"],
                "Destinataire_Nom": " ".join(r["consignee_name"]).strip(),
                "Destinataire_Adresse": addr_simple,
                "No_Conteneur": "; ".join(it["container_no"]),
                "No_Scelle": "; ".join(it["seal_no"]),
                "Numeros_Chassis": "; ".join(it["chassis"]),
                "Type_Colis": type_simple,
                "_cat_code": type_code,
                "Bebe_Au_Dos": bebe_au_dos,
                "Etat": statut,
                "Marque": marque,
                "Modele": modele,
                "Annee_Fabrication": annee_fab,
                "Couleur": couleur_bl,
                "Code_HS": code_hs_bl,
                "No_Moteur": no_moteur_bl,
                "LM": it["lm"] if it["lm"] is not None else 0.0,
                "Nb_Unites": it["qty"],
                "Poids_Kg":   it["weight"] if it["weight"] is not None else 0.0,
                "Tare_Kg":    it["tare"]   if it["tare"]   is not None else 0.0,
                "Volume_CBM": it["cbm"]    if it["cbm"]    is not None else 0.0,
                "Pays_Transit": transit_pays,
                "_transit_confiance": transit_conf,
            })

    df = pd.DataFrame(rows)
    if df.empty:
        return df
    # Regroupement : même B/L + même type (+ marque/modèle pour les véhicules)
    # → on additionne les quantités/poids. Les nouvelles colonnes (Port_Dechargement,
    # Nature_BL, Marque, Modele, Annee_Fabrication) doivent toutes figurer ici
    # pour ne pas être silencieusement supprimées par groupby.
    group_keys = [
        "Fichier", "Navire", "Voyage", "Port_Chargement", "Port_Dechargement",
        "BL_Numero", "Nature_BL",
        "Chargeur_Nom", "Destinataire_Nom", "Destinataire_Adresse",
        "Type_Colis", "_cat_code", "Bebe_Au_Dos", "Etat",
        "Marque", "Modele", "Annee_Fabrication",
        "Couleur", "Code_HS", "No_Moteur",
        "Pays_Transit", "_transit_confiance",
    ]
    agg = df.groupby(group_keys, dropna=False, sort=False).agg(
        No_Conteneur=("No_Conteneur", lambda s: "; ".join(dict.fromkeys(x for x in s if x))),
        No_Scelle=("No_Scelle", lambda s: "; ".join(dict.fromkeys(x for x in s if x))),
        Numeros_Chassis=("Numeros_Chassis", lambda s: "; ".join(dict.fromkeys(x for x in s if x))),
        Nb_Unites=("Nb_Unites", "sum"),
        Poids_Kg=("Poids_Kg", "sum"),
        Tare_Kg=("Tare_Kg", "sum"),
        Volume_CBM=("Volume_CBM", "sum"),
        LM=("LM", "sum"),
    ).reset_index()
    return agg


# Colonnes retenues par onglet (dans l'ordre d'affichage).
# Superset complet — l'agent peut masquer celles dont il n'a pas besoin
# depuis la page Structuration (profil par service ou sélection manuelle).
SHEET_COLUMNS = {
    "Vehicule": [
        "BL_Numero", "Nature_BL", "Navire", "Voyage",
        "Port_Chargement", "Port_Dechargement", "Pays_Transit",
        "Marque", "Modele", "Annee_Fabrication", "Couleur",
        "Numeros_Chassis", "No_Moteur", "Code_HS", "Etat",
        "Nb_Unites", "Poids_Kg", "Volume_CBM", "LM",
        "Chargeur_Nom", "Destinataire_Nom", "Destinataire_Adresse",
    ],
    "Conteneur": [
        "BL_Numero", "Nature_BL", "Navire", "Voyage",
        "Port_Chargement", "Port_Dechargement", "Pays_Transit",
        "No_Conteneur", "No_Scelle", "Type_Colis",
        "Nb_Unites", "Poids_Kg", "Tare_Kg", "Volume_CBM",
        "Chargeur_Nom", "Destinataire_Nom", "Destinataire_Adresse",
    ],
    "Colis": [
        "BL_Numero", "Nature_BL", "Navire", "Voyage",
        "Port_Chargement", "Port_Dechargement", "Pays_Transit",
        "Type_Colis", "Nb_Unites", "Poids_Kg", "Volume_CBM",
        "Chargeur_Nom", "Destinataire_Nom", "Destinataire_Adresse",
    ],
}
CAT_CODE_TO_SHEET = {"V": "Vehicule", "C": "Conteneur", "D": "Colis"}
# Noms d'onglets plus lisibles pour les agents (au lieu de "Detail_Cargaison_*")
SHEET_TAB_NAMES = {
    "Vehicule": "Cargaison groupée - Véhicules",
    "Conteneur": "Cargaison groupée - Conteneurs",
    "Colis": "Cargaison groupée - Colis",
}

# --- Fusion des onglets par catégorie (demande utilisateur 03/09) ---------
# Au lieu de 3 onglets détail + 3 onglets agrégés séparés, un seul onglet par
# niveau (Détail Cargaison / Cargaison groupée), avec une colonne Catégorie
# et une couleur de ligne par catégorie pour garder la lisibilité.
CATEGORY_LABELS = {"V": "Véhicules", "C": "Conteneurs", "D": "Colis"}
# Palette réutilisée du dashboard (bleu/orange/aqua, contraste + daltonisme
# déjà validés — voir CTX_CACHE_manifestes.md session 14/08). Teinte pastel
# (85% blanc + 15% couleur) pour le fond de ligne, couleur pleine pour le
# "chip" de la cellule Catégorie (texte blanc gras dessus).
CATEGORY_COLORS = {
    "Véhicules":  {"base": "2A78D6", "tint": "DFEBF9"},
    "Conteneurs": {"base": "EB6834", "tint": "FCE8E1"},
    "Colis":      {"base": "1BAF7A", "tint": "DDF3EB"},
}
# Ordre d'affichage préférentiel des colonnes de l'onglet "Cargaison groupée"
# fusionné — seules les colonnes réellement présentes (selon sheet_columns /
# profil d'affichage choisi par l'agent) sont gardées, dans cet ordre.
MERGED_AGGREGE_COLUMN_ORDER = [
    "BL_Numero", "Nature_BL", "Navire", "Voyage",
    "Port_Chargement", "Port_Dechargement", "Pays_Transit",
    "Marque", "Modele", "Annee_Fabrication", "Couleur",
    "Numeros_Chassis", "No_Moteur", "Code_HS", "Etat",
    "No_Conteneur", "No_Scelle", "Type_Colis",
    "Nb_Unites", "Poids_Kg", "Tare_Kg", "Volume_CBM", "LM",
    "Chargeur_Nom", "Destinataire_Nom", "Destinataire_Adresse",
]
# Colonnes fixes de l'onglet "Détail Cargaison" fusionné (pas de sélection
# par profil ici, comme avant la fusion — union des 3 anciens onglets détail).
MERGED_DETAIL_COLUMNS = [
    "Catégorie", "BL_Numero", "Nature_BL", "Navire", "Voyage",
    "Port_Chargement", "Port_Dechargement", "Pays_Transit",
    "Marque", "Modele", "Annee_Fabrication", "Chassis",
    "No_Conteneur", "No_Scelle",
    "Type_Colis", "N_Unite",
    "Etat", "Poids_Unitaire_Kg", "Tare_Kg", "Volume_CBM",
    "Chargeur_Nom", "Destinataire_Nom",
]

CARGO_TYPE_LABELS = {
    frozenset({"V"}): "🚗 Véhicules uniquement",
    frozenset({"C"}): "📦 Conteneurs uniquement",
    frozenset({"D"}): "📋 Colis uniquement",
    frozenset({"V", "C"}): "🔀 Mixte (Véhicules + Conteneurs)",
    frozenset({"V", "D"}): "🔀 Mixte (Véhicules + Colis)",
    frozenset({"C", "D"}): "🔀 Mixte (Conteneurs + Colis)",
    frozenset({"V", "C", "D"}): "🔀 Mixte (Véhicules + Conteneurs + Colis)",
}


def classify_cargo_type(df_subset):
    """Détermine le type de cargaison d'un navire/voyage à partir des
    catégories réellement présentes dans ses lignes ("Véhicules uniquement",
    "Mixte", ...). Certains manifestes ne listent que des véhicules, d'autres
    combinent plusieurs types — utile pour organiser l'aperçu, l'archive et
    la recherche sans que l'agent ait à ouvrir chaque fichier pour le savoir."""
    if df_subset.empty or "_cat_code" not in df_subset.columns:
        return "—"
    present = frozenset(df_subset["_cat_code"].dropna().unique())
    return CARGO_TYPE_LABELS.get(present, "—")


def records_to_item_dataframe(records):
    """Vue détaillée : 1 ligne = 1 item (type/quantité/poids/volume déjà
    correctement apparié), pour classification par tranche de volume
    (reproduit la logique du Tableau de classification des véhicules)."""
    rows = []
    for r in records:
        full_desc = " | ".join(dict.fromkeys(r["raw_desc_lines"]))
        for it in r["items"]:
            type_simple, type_code = simplify_type_colis(it["type_raw"])
            statut = item_status(it["type_raw"], full_desc)
            is_vehicle = type_code == "V"
            qty = it["qty"] or 1
            cbm = it["cbm"]
            avg_cbm = (cbm / qty) if (cbm is not None and qty) else None
            bucket = ""
            if is_vehicle and avg_cbm is not None:
                bucket = "<15m3" if avg_cbm < 15 else ("15-50m3" if avg_cbm <= 50 else ">50m3")
            rows.append({
                "Fichier": r["source_file"],
                "Navire_Voyage": r["vessel_voyage"],
                "Port_Chargement": r["port_of_loading"],
                "Origine_Cargaison": r["origin_port"] if r["origin_port"] != r["port_of_loading"] else "",
                "BL_Numero": r["bl_number"],
                "Type_Colis": type_simple,
                "Statut_Vehicule": statut,
                "Nb_Unites": qty,
                "Type_Colis_Texte": f'{it["qty"]}-{it["type_raw"]}',
                "Chassis": "; ".join(it["chassis"]),
                "Poids_Kg": it["weight"],
                "Volume_CBM": it["cbm"],
                "Tranche_Volume": bucket,
            })
    return pd.DataFrame(rows)


HEADER_FILL = "1F4E78"


def write_sheet(ws, data, title_lines=None):
    """Ecrit un DataFrame dans une feuille openpyxl, avec bandeau titre et
    en-tetes stylees."""
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    body_font = Font(name="Arial", size=10)
    title_font = Font(name="Arial", bold=True, size=11)

    if title_lines:
        for line in title_lines:
            ws.append([line])
            ws.cell(row=ws.max_row, column=1).font = title_font
        ws.append([])

    for row in dataframe_to_rows(data, index=False, header=True):
        ws.append(row)
    # openpyxl peut laisser une "ligne fantome" apres un append([]) vide ;
    # on deduit la position reelle de l'en-tete a partir du nombre de
    # lignes de donnees effectivement ecrites (fiable), pas d'un compteur
    # de lignes calcule avant l'ecriture.
    header_row_idx = ws.max_row - len(data)
    for j, col in enumerate(data.columns, start=1):
        cell = ws.cell(row=header_row_idx, column=j)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=header_row_idx + 1):
        for cell in row:
            cell.font = body_font
    ws.freeze_panes = f"A{header_row_idx + 1}"
    for i, col in enumerate(data.columns, start=1):
        col_lens = data[col].map(lambda v: len(str(v)) if pd.notna(v) else 0)
        maxlen = max(len(str(col)), col_lens.max() if len(data) else 0)
        ws.column_dimensions[get_column_letter(i)].width = min(max(maxlen + 2, 10), 45)
    # Active le filtre automatique Excel sur la ligne d'en-tete, pour que les
    # agents puissent filtrer/trier chaque colonne directement dans Excel.
    last_col_letter = get_column_letter(len(data.columns))
    last_row = header_row_idx + max(len(data), 1)
    ws.auto_filter.ref = f"A{header_row_idx}:{last_col_letter}{last_row}"
    return header_row_idx


def _apply_category_colors(ws, data, category_col, header_row_idx):
    """Colore chaque ligne de données selon sa catégorie (teinte pastel sur
    toute la ligne + "chip" plein sur la cellule Catégorie), pour que les
    3 blocs Véhicules/Conteneurs/Colis restent identifiables au premier
    coup d'oeil dans l'onglet fusionné."""
    from openpyxl.styles import PatternFill, Font

    if category_col not in data.columns or data.empty:
        return
    cat_col_idx = list(data.columns).index(category_col) + 1
    chip_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    for i, cat in enumerate(data[category_col].tolist()):
        colors = CATEGORY_COLORS.get(cat)
        if not colors:
            continue
        excel_row = header_row_idx + 1 + i
        tint_fill = PatternFill(start_color=colors["tint"], end_color=colors["tint"], fill_type="solid")
        for cell in ws[excel_row]:
            cell.fill = tint_fill
        chip_cell = ws.cell(row=excel_row, column=cat_col_idx)
        chip_cell.fill = PatternFill(start_color=colors["base"], end_color=colors["base"], fill_type="solid")
        chip_cell.font = chip_font


def _rows_vehicule_detail(g_bl):
    """Calcule les lignes détail Véhicule (1 ligne par numéro de châssis
    identifié), sans les écrire dans une feuille. Liste vide si aucun
    chassis n'est présent dans ce manifeste (manifeste conteneur-only ou
    chassis non listés dans le PDF source)."""
    g_veh = g_bl[g_bl["_cat_code"] == "V"].copy()
    if g_veh.empty:
        return []
    mask = g_veh["Numeros_Chassis"].fillna("").astype(str).str.strip() != ""
    g_veh = g_veh[mask]
    if g_veh.empty:
        return []

    rows = []
    for _, r in g_veh.iterrows():
        chassis_list = [c.strip() for c in str(r["Numeros_Chassis"]).split(";") if c.strip()]
        if not chassis_list:
            continue
        nb = max(int(r.get("Nb_Unites") or 1), 1)
        poids_kg = r.get("Poids_Kg")
        # Un "bebe au dos" (remorque/vehicule empile) a delibirement Poids_Kg=0
        # (non fourni separement dans le PDF, deja compte sur l'unite
        # principale) — le traiter comme "0.0 kg confirme" tromperait l'agent.
        # Poids_Unitaire_Kg reste vide (None) dans ce cas plutot qu'un faux 0.
        try:
            poids_unit = round(float(poids_kg) / nb, 1) if poids_kg else None
        except (TypeError, ValueError, ZeroDivisionError):
            poids_unit = None
        # Volume par unité (même logique que le poids ci-dessus) — clé
        # "Volume_CBM" (même nom que pour Conteneur/Colis, déjà des valeurs
        # par-unité) pour être exposé dans l'onglet "Détail Cargaison"
        # (MERGED_DETAIL_COLUMNS) ET réutilisé par l'onglet "Pré-Masque
        # complet" (classification C/V/T) et la classification véhicules
        # par POL (tâche 11a, calculée à la volée depuis cet onglet).
        cbm = r.get("Volume_CBM")
        try:
            volume_unit = round(float(cbm) / nb, 3) if cbm else None
        except (TypeError, ValueError, ZeroDivisionError):
            volume_unit = None
        for ch in chassis_list:
            rows.append({
                "BL_Numero":           r.get("BL_Numero", ""),
                "Nature_BL":           r.get("Nature_BL", ""),
                "Navire":              r.get("Navire", ""),
                "Voyage":              r.get("Voyage", ""),
                "Port_Chargement":     r.get("Port_Chargement", ""),
                "Port_Dechargement":   r.get("Port_Dechargement", ""),
                "Pays_Transit":        r.get("Pays_Transit", ""),
                "Marque":              r.get("Marque", ""),
                "Modele":              r.get("Modele", ""),
                "Annee_Fabrication":   r.get("Annee_Fabrication", ""),
                "Chassis":             ch,
                "Etat":                r.get("Etat", ""),
                "Poids_Unitaire_Kg":   poids_unit,
                "Volume_CBM":          volume_unit,
                "Chargeur_Nom":        r.get("Chargeur_Nom", ""),
                "Destinataire_Nom":    r.get("Destinataire_Nom", ""),
            })
    return rows


def _build_bebe_au_dos_sheet(wb, g_bl, title_lines):
    """Ajoute l'onglet dédié "Bébé au dos" — une ligne par matricule d'engin
    porté (remorque attelée / véhicule empilé), isolé des onglets véhicules
    standards. Décision (02/09, retour utilisateur) : la colonne Bebe_Au_Dos
    est une info de VÉRIFICATION pure pour l'agent, pas une donnée qui existe
    dans le fichier de référence utilisé habituellement — on la sort donc des
    onglets Véhicule/Détail Chassis (qui restent identiques à la structure
    connue des agents) et on centralise ici tous les engins portés, avec leur
    B/L et châssis, pour vérification/complément ciblé.
    Absent si aucun bébé au dos n'est présent dans ce manifeste."""
    g_veh = g_bl[(g_bl["_cat_code"] == "V") & (g_bl.get("Bebe_Au_Dos", "") == "Oui")].copy()
    if g_veh.empty:
        return
    mask = g_veh["Numeros_Chassis"].fillna("").astype(str).str.strip() != ""
    g_veh = g_veh[mask]
    if g_veh.empty:
        return

    rows = []
    for _, r in g_veh.iterrows():
        chassis_list = [c.strip() for c in str(r["Numeros_Chassis"]).split(";") if c.strip()]
        for ch in chassis_list:
            rows.append({
                "BL_Numero":         r.get("BL_Numero", ""),
                "Navire":            r.get("Navire", ""),
                "Voyage":            r.get("Voyage", ""),
                "Chassis":           ch,
                "Marque":            r.get("Marque", ""),
                "Modele":            r.get("Modele", ""),
                "Etat":              r.get("Etat", ""),
                "Chargeur_Nom":      r.get("Chargeur_Nom", ""),
                "Destinataire_Nom":  r.get("Destinataire_Nom", ""),
            })
    if not rows:
        return
    df_bb = pd.DataFrame(rows)
    ws = wb.create_sheet("Bébé au dos (engins portés)")
    write_sheet(ws, df_bb, title_lines=title_lines)


def _rows_conteneur_detail(g_bl):
    """Calcule les lignes détail Conteneur (1 ligne par numéro de conteneur),
    sans les écrire dans une feuille. Liste vide si aucun conteneur n'est
    présent ou si aucun numéro n'est renseigné."""
    g_cont = g_bl[g_bl["_cat_code"] == "C"].copy()
    if g_cont.empty:
        return []
    mask = g_cont["No_Conteneur"].fillna("").astype(str).str.strip() != ""
    g_cont = g_cont[mask]
    if g_cont.empty:
        return []

    rows = []
    for _, r in g_cont.iterrows():
        cont_list = [c.strip() for c in str(r["No_Conteneur"]).split(";") if c.strip()]
        seal_list = [s.strip() for s in str(r.get("No_Scelle", "")).split(";") if s.strip()]
        if not cont_list:
            continue
        nb = max(int(r.get("Nb_Unites") or 1), 1)
        poids_kg = r.get("Poids_Kg")
        tare_kg  = r.get("Tare_Kg")
        cbm      = r.get("Volume_CBM")
        try:
            poids_unit = round(float(poids_kg) / nb, 1) if poids_kg else None
        except (TypeError, ValueError, ZeroDivisionError):
            poids_unit = None
        try:
            tare_unit = round(float(tare_kg) / nb, 1) if tare_kg else None
        except (TypeError, ValueError, ZeroDivisionError):
            tare_unit = None
        try:
            cbm_unit = round(float(cbm) / nb, 3) if cbm else None
        except (TypeError, ValueError, ZeroDivisionError):
            cbm_unit = None
        for i, cont in enumerate(cont_list):
            seal = seal_list[i] if i < len(seal_list) else ""
            rows.append({
                "BL_Numero":         r.get("BL_Numero", ""),
                "Nature_BL":         r.get("Nature_BL", ""),
                "Navire":            r.get("Navire", ""),
                "Voyage":            r.get("Voyage", ""),
                "Port_Chargement":   r.get("Port_Chargement", ""),
                "Port_Dechargement": r.get("Port_Dechargement", ""),
                "Pays_Transit":      r.get("Pays_Transit", ""),
                "No_Conteneur":      cont,
                "No_Scelle":         seal,
                "Type_Colis":        r.get("Type_Colis", ""),
                "Poids_Unitaire_Kg": poids_unit,
                "Tare_Kg":           tare_unit,
                "Volume_CBM":        cbm_unit,
                "Chargeur_Nom":      r.get("Chargeur_Nom", ""),
                "Destinataire_Nom":  r.get("Destinataire_Nom", ""),
            })
    return rows


def _rows_colis_detail(g_bl):
    """Calcule les lignes détail Colis (1 ligne par unité, expansion par
    Nb_Unites), sans les écrire dans une feuille. Liste vide si aucun colis
    n'est présent dans ce manifeste."""
    g_colis = g_bl[g_bl["_cat_code"] == "D"].copy()
    if g_colis.empty:
        return []

    rows = []
    for _, r in g_colis.iterrows():
        nb = max(int(r.get("Nb_Unites") or 1), 1)
        poids_kg = r.get("Poids_Kg")
        cbm      = r.get("Volume_CBM")
        try:
            poids_unit = round(float(poids_kg) / nb, 1) if poids_kg else None
        except (TypeError, ValueError, ZeroDivisionError):
            poids_unit = None
        try:
            cbm_unit = round(float(cbm) / nb, 3) if cbm else None
        except (TypeError, ValueError, ZeroDivisionError):
            cbm_unit = None
        for i in range(nb):
            rows.append({
                "BL_Numero":         r.get("BL_Numero", ""),
                "Nature_BL":         r.get("Nature_BL", ""),
                "Navire":            r.get("Navire", ""),
                "Voyage":            r.get("Voyage", ""),
                "Port_Chargement":   r.get("Port_Chargement", ""),
                "Port_Dechargement": r.get("Port_Dechargement", ""),
                "Pays_Transit":      r.get("Pays_Transit", ""),
                "Type_Colis":        r.get("Type_Colis", ""),
                "N_Unite":           i + 1,
                "Poids_Unitaire_Kg": poids_unit,
                "Volume_CBM":        cbm_unit,
                "Chargeur_Nom":      r.get("Chargeur_Nom", ""),
                "Destinataire_Nom":  r.get("Destinataire_Nom", ""),
            })
    return rows


# Colonnes de l'onglet "Pré-Masque complet" — mêmes noms et même ordre que le
# pré-masque IPAKI/TETRAX généré depuis un manifest navire à grue (voir
# crane_manifest_parser.generate_premasque_excel), pour que l'agent retrouve
# une structure identique quelle que soit la source. Onglet VÉHICULES
# UNIQUEMENT (une ligne par châssis) : les colonnes ci-dessous (CHÂSSIS,
# MARQUE, TYPE/TAILLE...) n'ont pas d'équivalent pertinent pour les
# conteneurs/colis, qui restent couverts par les onglets Détail Cargaison /
# Cargaison groupée. Colonnes absentes du manifeste PDF source (booking :
# MODE DE TRANSPORT, ESCALE TETRAX/IPAKI, POL/POD IPAKI, CLIENT — distinct du
# destinataire) restent volontairement vides, à compléter par les agents —
# même logique que les colonnes de booking de l'onglet Reporting.
PREMASQUE_COMPLET_COLUMNS = [
    "NBRE", "MODE DE TRANSPORT", "ESCALE TETRAX", "NATURE BL",
    "POL TETRAX", "POD TETRAX", "FINAL DESTINATION TETRAX",
    "POIDS TETRAX (KG)", "TYPE / TAILLE", "VOLUME TETRAX", "BL",
    "ESCALE IPAKI", "POL IPAKI", "POD IPAKI", "FINAL DESTINATION IPAKI",
    "MARQUE", "MODELE", "MARQUE & MODELE", "ETAT", "ANNEE DE FABRICATION",
    "CHÂSSIS", "TYPE D'ACTION", "POIDS IPAKI (TON)", "VOLUME",
    "BLItem YardItemCode", "CLIENT", "OBSERVATION",
]


def _volume_tranche(volume_cbm):
    """Classification C (<15m³) / V (15-50m³) / T (>50m³) — même seuils que
    la colonne "TYPE / TAILLE" du pré-masque IPAKI (crane_manifest_parser) et
    que Tranche_Volume utilisé pour la classification véhicules. "" si volume
    inconnu (ne pas classifier arbitrairement)."""
    if volume_cbm is None:
        return ""
    try:
        v = float(volume_cbm)
    except (TypeError, ValueError):
        return ""
    if v != v:  # NaN (pd.NA/np.nan passent float() sans lever — jamais classer un volume manquant)
        return ""
    if v <= 0:
        return ""
    if v < 15:
        return "C"
    if v <= 50:
        return "V"
    return "T"


def _rows_premasque_complet(g_bl):
    """Calcule les lignes de l'onglet "Pré-Masque complet" (1 ligne par
    châssis, véhicules uniquement) à partir des lignes détail Véhicule déjà
    calculées par _rows_vehicule_detail — évite de dupliquer la logique
    d'expansion par châssis / calcul du poids et volume unitaires."""
    veh_rows = _rows_vehicule_detail(g_bl)
    if not veh_rows:
        return []

    rows = []
    for i, r in enumerate(veh_rows, start=1):
        nature_bl = r.get("Nature_BL", "") or "Import"
        volume_unit = r.get("Volume_CBM")
        taille = _volume_tranche(volume_unit)
        poids_unit = r.get("Poids_Unitaire_Kg")
        marque = r.get("Marque", "")
        modele = r.get("Modele", "")
        marque_modele = f"{marque} {modele}".strip() if marque else modele
        dest_finale = r.get("Pays_Transit") or r.get("Port_Dechargement", "")
        type_action = {"Export": "EXPORT", "Transbo": "TRANSBO"}.get(nature_bl, "IMPORT")
        observation = f"TRANSIT VERS {r['Pays_Transit']}" if r.get("Pays_Transit") else ""

        rows.append({
            "NBRE":                     i,
            "MODE DE TRANSPORT":        "",
            "ESCALE TETRAX":            "",
            "NATURE BL":                nature_bl,
            "POL TETRAX":               r.get("Port_Chargement", ""),
            "POD TETRAX":               r.get("Port_Dechargement", ""),
            "FINAL DESTINATION TETRAX": dest_finale,
            "POIDS TETRAX (KG)":        int(poids_unit) if poids_unit else "",
            "TYPE / TAILLE":            taille,
            "VOLUME TETRAX":            volume_unit if volume_unit else "",
            "BL":                       r.get("BL_Numero", ""),
            "ESCALE IPAKI":             "",
            "POL IPAKI":                "",
            "POD IPAKI":                "",
            "FINAL DESTINATION IPAKI":  dest_finale,
            "MARQUE":                   marque,
            "MODELE":                   modele,
            "MARQUE & MODELE":          marque_modele,
            "ETAT":                     r.get("Etat", ""),
            "ANNEE DE FABRICATION":     r.get("Annee_Fabrication", ""),
            "CHÂSSIS":                  r.get("Chassis", ""),
            "TYPE D'ACTION":            type_action,
            "POIDS IPAKI (TON)":        round(poids_unit / 1000, 3) if poids_unit else "",
            "VOLUME":                   volume_unit if volume_unit else "",
            "BLItem YardItemCode":      f"VEH {'< 15m3' if taille == 'C' else '> 15m3'}" if taille else "",
            "CLIENT":                   "",
            "OBSERVATION":              observation,
        })
    return rows


def _build_premasque_complet_sheet(wb, g_bl, title_lines):
    """Ajoute l'onglet "Pré-Masque complet" (structure TETRAX/IPAKI, véhicules
    uniquement). Absent si le manifeste ne contient aucun véhicule avec
    châssis identifié (ex. manifeste conteneur-only)."""
    rows = _rows_premasque_complet(g_bl)
    if not rows:
        return
    df = pd.DataFrame(rows).reindex(columns=PREMASQUE_COMPLET_COLUMNS)
    ws = wb.create_sheet("Pré-Masque complet")
    write_sheet(ws, df, title_lines=title_lines)


def _build_detail_sheet_merged(wb, g_bl, title_lines):
    """Ajoute l'onglet fusionné "Détail Cargaison" — union des 3 anciens
    onglets détail (Véhicules par châssis / Conteneurs par numéro / Colis
    par unité), lignes groupées par catégorie (Véhicules puis Conteneurs
    puis Colis) et colorées (voir CATEGORY_COLORS). Absent si aucune donnée
    détail dans aucune des 3 catégories."""
    rows = []
    for label, row_list in (
        (CATEGORY_LABELS["V"], _rows_vehicule_detail(g_bl)),
        (CATEGORY_LABELS["C"], _rows_conteneur_detail(g_bl)),
        (CATEGORY_LABELS["D"], _rows_colis_detail(g_bl)),
    ):
        for r in row_list:
            rows.append({"Catégorie": label, **r})
    if not rows:
        return
    df = pd.DataFrame(rows).reindex(columns=MERGED_DETAIL_COLUMNS)
    ws = wb.create_sheet("Détail Cargaison")
    header_row_idx = write_sheet(ws, df, title_lines=title_lines)
    _apply_category_colors(ws, df, "Catégorie", header_row_idx)


def _build_aggrege_sheet_merged(wb, g_bl, title_lines, cols_map):
    """Ajoute l'onglet fusionné "Cargaison groupée" — union des 3 anciens
    onglets agrégés (structure groupée par B/L), lignes groupées par
    catégorie et colorées. Respecte cols_map (sélection de colonnes par
    profil Reporting/Opérations/Personnalisé) : seules les colonnes
    effectivement sélectionnées pour au moins une catégorie apparaissent.
    Absent si aucune donnée dans aucune des 3 catégories."""
    frames = []
    present_cols = set()
    for cat_code, sheet_name in CAT_CODE_TO_SHEET.items():
        g_cat = g_bl[g_bl["_cat_code"] == cat_code]
        if g_cat.empty:
            continue
        cols = [c for c in cols_map.get(sheet_name, []) if c in g_cat.columns]
        if not cols:
            continue
        sub = g_cat[cols].reset_index(drop=True)
        sub.insert(0, "Catégorie", CATEGORY_LABELS[cat_code])
        present_cols.update(cols)
        frames.append(sub)
    if not frames:
        return
    ordered_cols = ["Catégorie"] + [c for c in MERGED_AGGREGE_COLUMN_ORDER if c in present_cols]
    # Filet de sécurité si une colonne future n'est pas dans l'ordre préférentiel ci-dessus.
    ordered_cols += [c for c in present_cols if c not in ordered_cols]
    df = pd.concat(frames, ignore_index=True).reindex(columns=ordered_cols)
    ws = wb.create_sheet("Cargaison groupée")
    header_row_idx = write_sheet(ws, df, title_lines=title_lines)
    _apply_category_colors(ws, df, "Catégorie", header_row_idx)


def build_workbook_bytes(g_bl, navire, voyage, sheet_columns=None):
    """Construit un classeur Excel pour UN navire/voyage deja filtre.

    Onglets générés :
      - Pré-Masque complet (structure TETRAX/IPAKI, véhicules uniquement —
        une ligne par châssis) : mêmes colonnes que le pré-masque généré
        depuis un manifest navire à grue, pour que l'agent retrouve une
        structure identique quelle que soit la source. Colonnes booking sans
        équivalent dans le PDF (MODE DE TRANSPORT, ESCALE TETRAX/IPAKI,
        POL/POD IPAKI, CLIENT) restent vides, à compléter par les agents.
        Absent si aucun véhicule avec châssis identifié.
      - Détail Cargaison (Véhicules/Conteneurs/Colis fusionnés, groupés par
        catégorie et colorés — une ligne par unité individuelle : châssis,
        numéro de conteneur, ou unité de colis)
      - Bébé au dos (engins portés) — onglet dédié, uniquement si présent ;
        reste séparé du Détail Cargaison (pas de colonne Bebe_Au_Dos dans
        celui-ci, comportement inchangé depuis le 02/09)
      - Cargaison groupée (Véhicules/Conteneurs/Colis fusionnés, structure
        groupée par B/L, groupée par catégorie et colorée)

    sheet_columns permet de surcharger les colonnes visibles par catégorie
    dans l'onglet Cargaison groupée (ex : choix de l'agent dans l'interface,
    profil Reporting/Opérations/Personnalisé)."""
    from openpyxl import Workbook
    import io

    cols_map = sheet_columns if sheet_columns is not None else SHEET_COLUMNS
    title = [f"Navire : {navire}", f"Voyage : {voyage}", "Port de déchargement : ABIDJAN"]

    wb = Workbook()
    wb.remove(wb.active)  # feuille par defaut vide — les onglets ci-dessous la remplacent

    _build_premasque_complet_sheet(wb, g_bl, title)          # structure Pré-Masque TETRAX/IPAKI (véhicules)
    _build_detail_sheet_merged(wb, g_bl, title)              # détail unifié, par catégorie colorée
    _build_bebe_au_dos_sheet(wb, g_bl, title)                # engins portés, isolés (inchangé)
    _build_aggrege_sheet_merged(wb, g_bl, title, cols_map)    # agrégé unifié, par catégorie colorée

    if not wb.sheetnames:
        # Cas limite : aucune donnee du tout — garder un classeur valide.
        wb.create_sheet("Détail Cargaison")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_excel_per_vessel(df_bl, output_dir):
    """Génère un classeur Excel par Navire/Voyage sur disque (usage CLI/local)."""
    import os
    os.makedirs(output_dir, exist_ok=True)
    written = []
    for (navire, voyage), g_bl in df_bl.groupby(["Navire", "Voyage"]):
        safe_name = re.sub(r'[^A-Za-z0-9]+', '_', f"{navire}_{voyage}").strip('_')
        buf = build_workbook_bytes(g_bl, navire, voyage)
        path = os.path.join(output_dir, f"Manifeste_{safe_name}.xlsx")
        with open(path, "wb") as f:
            f.write(buf.getbuffer())
        written.append(path)
    return written


if __name__ == "__main__":
    files = [
        ("/mnt/user-data/uploads/GSE0426_-_DAKAR_FINAL.pdf", "GSE0426_-_DAKAR_FINAL.pdf"),
        ("/mnt/user-data/uploads/GSE0426_-_GIOIA_TAURO.pdf", "GSE0426_-_GIOIA_TAURO.pdf"),
    ]
    all_records = []
    for path, label in files:
        recs = parse_manifest(path, label)
        print(f"{label}: {len(recs)} B/L extraits")
        all_records.extend(recs)

    df_bl = records_to_dataframe(all_records)

    outputs = build_excel_per_vessel(df_bl, "/home/claude/work/out")
    print("\nFichiers générés:")
    for p in outputs:
        print(" -", p)
    print(f"\nTotal lignes groupées (BL x Type): {len(df_bl)}")
    print("\nRépartition par catégorie:")
    print(df_bl["_cat_code"].value_counts())
    print("\nRépartition Etat:")
    print(df_bl["Etat"].value_counts(dropna=False))
