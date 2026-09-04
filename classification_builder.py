"""
classification_builder.py — Tableau de classification des conteneurs par
POL (port de chargement) et tranche de volume (<15m³ / 15-50m³ / >50m³).

Contexte (voir claude/ANALYSE_TABLEAU_CLASSIFICATION_VEHICULES_2026-09-03.md
dans le projet Claude) : remplace le fichier manuel à ~150 onglets (un par
escale) par un calcul automatique depuis les manifestes déjà structurés.

Repositionné le 04/09 (retour utilisateur, sur échantillon réel de 13
manifestes) : la quasi-totalité des manifestes Grimaldi traités par TERRA
sont des manifestes conteneurs/groupage, pas des manifestes RORO châssis —
la classification porte donc sur l'onglet "Cargaison groupée - Conteneurs"
(1 ligne = 1 conteneur physique), pas sur les véhicules. Le principe reste
inchangé : POL en lignes, tranche de volume en colonnes (NOMBRE/TONNAGE/
VOLUME par tranche), classée d'après Volume_CBM (déjà renseigné par
conteneur, capacité standard 20/40 pieds) — même seuils <15/15-50/>50m3 que
la version précédente. Le tableau reste agrégé (une somme par POL, pas une
ligne par conteneur) : voir pivot_pol_tranche.

Décisions MVP :
  - 04/09 (retour utilisateur) : TOUS les B/L sont classifiés, quel que soit
    Nature_BL (Import/Export/Transb.) — plus d'exclusion. Le détail par
    nature reste visible dans le diagnostic (diag["par_nature"]).
  - Calcul à la volée depuis les exports déjà archivés (même principe que
    reporting_builder.fetch_voyage_detail), pas de nouvelle table de
    données brutes.
"""
import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import reporting_builder as rbld
from manifest_parser import _volume_tranche, HEADER_FILL

TRANCHES = ["C", "V", "T"]  # _volume_tranche : C=<15m3, V=15-50m3, T=>50m3
# Libellés "modèle" (majuscules, comme le fichier de référence x150 onglets)
# utilisés pour l'export Excel ET l'affichage écran — symbole m³ pour la
# lisibilité (demande utilisateur 04/09).
TRANCHE_LABELS = {"C": "CONTENEUR < 15 M³", "V": "CONTENEUR 15-50 M³", "T": "CONTENEUR > 50 M³"}
# Sous-colonnes affichées à l'écran (Streamlit) — mêmes données que NOMBRE/
# TONNAGE/VOLUME (clés internes, utilisées par pivot_pol_tranche/Excel) mais
# avec unité explicite pour l'agent qui regarde juste le tableau.
SUB_LABELS_DISPLAY = {"NOMBRE": "Nombre", "TONNAGE": "Tonnage (t)", "VOLUME": "Volume (m³)"}

_EMPTY_COLS = ["POL", "Tranche", "Poids_Unitaire_Kg", "Volume_CBM"]


def classify_conteneurs(navire: str, voyage: str):
    """Retourne (df_classifie, diag) pour un Navire/Voyage donné.

    df_classifie : 1 ligne par conteneur physique, colonnes POL
    (Port_Chargement), Tranche (<15m3/15-50m3/>50m3, vide si volume inconnu),
    Poids_Unitaire_Kg, Volume_CBM. Vide si aucun conteneur classifiable.

    diag : {total_conteneurs, sans_volume (volume manquant — ré-traitement
    nécessaire), par_nature ({"Import": n, "Transb.": n, ...} — répartition
    informative, plus aucune exclusion depuis le 04/09)} — en plus du
    diagnostic de fetch_voyage_detail (accessible séparément via
    reporting_builder.fetch_voyage_detail si besoin d'un détail plus fin sur
    les échecs de lecture d'export)."""
    result, _used_df, _ports, _fetch_diag = rbld.fetch_voyage_detail(navire, voyage)
    df_cont = result.get("Conteneur", pd.DataFrame())
    diag = {"total_conteneurs": len(df_cont), "sans_volume": 0, "par_nature": {}}
    if df_cont.empty:
        return pd.DataFrame(columns=_EMPTY_COLS), diag

    # rbld._col(df, name) : renvoie df[name] si présente, sinon une Série de
    # valeurs par défaut de la bonne longueur/index — contrairement à
    # df.get(name) (ou df.get(name, default) avec un default scalaire), qui
    # renvoie None/le scalaire brut (PAS une Série) quand la colonne est
    # absente (bug réel constaté en production, 03/09).
    nature = rbld._col(df_cont, "Nature_BL").astype(str).str.strip()
    nature_label = nature.replace("", "(non renseigné)")
    diag["par_nature"] = {k: int(v) for k, v in nature_label.value_counts().items()}

    volume = pd.to_numeric(rbld._col(df_cont, "Volume_CBM"), errors="coerce")
    tranche = volume.map(_volume_tranche)
    diag["sans_volume"] = int((tranche == "").sum())
    poids = pd.to_numeric(rbld._col(df_cont, "Poids_Unitaire_Kg"), errors="coerce")

    df_out = pd.DataFrame({
        "POL": rbld._col(df_cont, "Port_Chargement").astype(str).str.strip(),
        "Tranche": tranche,
        "Poids_Unitaire_Kg": poids,
        "Volume_CBM": volume,
    })
    return df_out, diag


def pivot_pol_tranche(df_classifie: pd.DataFrame) -> pd.DataFrame:
    """Construit le tableau croisé POL (lignes, avec ligne TOTAL finale) ×
    tranche de volume (groupes de colonnes NOMBRE/TONNAGE/VOLUME), même
    structure que le fichier de référence (colonnes multi-niveaux aplaties en
    "<groupe> - <sous-colonne>" pour rester un DataFrame simple ; l'export
    Excel reconstruira l'en-tête à 2 niveaux visuel). Un ensemble agrégé —
    une somme par POL (nombre de conteneurs + tonnage + volume), pas une
    ligne par conteneur (voir demande utilisateur 04/09). Lignes sans Tranche
    (volume manquant) exclues du croisé mais comptées dans le diagnostic de
    classify_conteneurs — jamais classées au hasard."""
    cols = ["POL"] + [f"{TRANCHE_LABELS[t]} - {sub}" for t in TRANCHES for sub in ("NOMBRE", "TONNAGE", "VOLUME")]
    if df_classifie.empty:
        return pd.DataFrame(columns=cols)

    df = df_classifie[df_classifie["Tranche"] != ""].copy()
    if df.empty:
        return pd.DataFrame(columns=cols)

    rows = []
    for pol, g in df.groupby("POL", sort=True):
        row = {"POL": pol}
        for t in TRANCHES:
            gt = g[g["Tranche"] == t]
            row[f"{TRANCHE_LABELS[t]} - NOMBRE"] = len(gt)
            row[f"{TRANCHE_LABELS[t]} - TONNAGE"] = round(gt["Poids_Unitaire_Kg"].sum() / 1000.0, 3) if len(gt) else 0
            row[f"{TRANCHE_LABELS[t]} - VOLUME"] = round(gt["Volume_CBM"].sum(), 2) if len(gt) else 0
        rows.append(row)

    total = {"POL": "TOTAL"}
    for t in TRANCHES:
        total[f"{TRANCHE_LABELS[t]} - NOMBRE"] = sum(r[f"{TRANCHE_LABELS[t]} - NOMBRE"] for r in rows)
        total[f"{TRANCHE_LABELS[t]} - TONNAGE"] = round(sum(r[f"{TRANCHE_LABELS[t]} - TONNAGE"] for r in rows), 3)
        total[f"{TRANCHE_LABELS[t]} - VOLUME"] = round(sum(r[f"{TRANCHE_LABELS[t]} - VOLUME"] for r in rows), 2)
    rows.append(total)

    return pd.DataFrame(rows, columns=cols)


def pivot_pol_tranche_styled(df_classifie: pd.DataFrame):
    """Version "présentable" de pivot_pol_tranche pour l'affichage écran
    (st.dataframe) — mêmes chiffres, mais :
      - POL en index (plus de colonne technique "POL" collée aux nombres) ;
      - en-tête à 2 niveaux (groupe de tranche / Nombre-Tonnage-Volume),
        rendu par Streamlit comme des colonnes fusionnées — plus proche du
        fichier de référence qu'une colonne plate "CONTENEUR < 15 M³ - TONNAGE" ;
      - nombres formatés (séparateur de milliers, décimales adaptées à
        l'unité) plutôt que des flottants bruts ;
      - ligne TOTAL mise en évidence (fond bleu foncé, texte blanc, gras).
    Retourne un pandas Styler (accepté directement par st.dataframe) ; None
    si rien à afficher (le pivot croisé est vide)."""
    pivot = pivot_pol_tranche(df_classifie)
    if pivot.empty:
        return None

    df = pivot.set_index("POL")
    df.columns = pd.MultiIndex.from_tuples(
        [tuple(c.split(" - ")) for c in df.columns],
    )
    df = df.rename(columns=SUB_LABELS_DISPLAY, level=1)

    fmt = {}
    for t in TRANCHES:
        fmt[(TRANCHE_LABELS[t], "Nombre")] = "{:,.0f}"
        fmt[(TRANCHE_LABELS[t], "Tonnage (t)")] = "{:,.3f}"
        fmt[(TRANCHE_LABELS[t], "Volume (m³)")] = "{:,.2f}"

    def _highlight_total(row):
        is_total = row.name == "TOTAL"
        return ["background-color: #1F4E78; color: white; font-weight: bold;" if is_total else ""] * len(row)

    styler = (
        df.style
        .format(fmt, na_rep="—")
        .apply(_highlight_total, axis=1)
        .set_table_styles([
            {"selector": "th", "props": [("text-align", "center"), ("font-weight", "bold")]},
        ])
    )
    return styler


# ---------------------------------------------------------------------------
# Export Excel — mise en page fidèle au fichier de référence x150 onglets
# (tâche 11d, voir claude/ANALYSE_TABLEAU_CLASSIFICATION_VEHICULES_2026-09-03.md
# section 1) : titre, en-tête à 2 niveaux (3 groupes de tranche × 3
# sous-colonnes NOMBRE/TONNAGE/VOLUME), une ligne PAR POL — un ensemble
# agrégé (nombre de conteneurs + tonnage + volume cumulés), pas une ligne par
# conteneur (demande utilisateur 04/09) — puis une ligne TOTAL en bas.
# Repose directement sur pivot_pol_tranche (même chiffres que l'affichage
# écran).
# ---------------------------------------------------------------------------
TOTAL_FILL = "1F4E78"     # même bleu que l'en-tête — ligne TOTAL générale
SUBTITLE_FILL = "DFEBF9"  # bleu très pastel — bandeau Navire/Voyage/Escale
ZEBRA_FILL = "F5F8FC"     # gris-bleu très clair — une ligne POL sur deux
BORDER_COLOR = "B8C4D9"


def build_classification_workbook_bytes(df_classifie: pd.DataFrame, navire: str, voyage: str,
                                          escale_info: dict | None = None) -> io.BytesIO:
    """escale_info optionnel : {"date_escale"} (voir tracking.get_suivi_escale)
    — affiché en bandeau titre si fourni, purement informatif, n'affecte pas
    le calcul.

    Mise en page (revue 04/09, demande utilisateur "plus lisible, plus
    agréable visuellement") : titre fusionné et centré sur toute la largeur
    du tableau (au lieu d'une simple cellule A1), bandeau Navire/Voyage/
    Escale sur fond pastel, quadrillage fin sur l'ensemble du tableau,
    lignes POL alternées (zébrage léger) pour suivre une ligne à l'oeil,
    nombres formatés (séparateur de milliers, décimales adaptées à
    l'unité — entier pour NOMBRE, 3 décimales pour TONNAGE, 2 pour VOLUME)."""
    n_cols = 1 + len(TRANCHES) * 3  # POL + 3 tranches × 3 sous-colonnes

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    title_font = Font(name="Arial", bold=True, size=14, color="1F4E78")
    subtitle_font = Font(name="Arial", bold=True, size=10, color="1F4E78")
    subtitle_fill = PatternFill(start_color=SUBTITLE_FILL, end_color=SUBTITLE_FILL, fill_type="solid")
    body_font = Font(name="Arial", size=10)
    pol_font = Font(name="Arial", size=10, bold=True)
    total_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    total_fill = PatternFill(start_color=TOTAL_FILL, end_color=TOTAL_FILL, fill_type="solid")
    zebra_fill = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Classification"

    # ── Bandeau titre : fusionné et centré sur toute la largeur du tableau,
    # gros caractère, façon en-tête de rapport plutôt qu'une simple cellule
    # A1 en haut à gauche ──
    ws.append(["TABLEAU DE CLASSIFICATION DES CONTENEURS PAR POL ET VOLUME"])
    title_row = ws.max_row
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=n_cols)
    title_cell = ws.cell(row=title_row, column=1)
    title_cell.font = title_font
    title_cell.alignment = center
    ws.row_dimensions[title_row].height = 24

    # ── Sous-bandeau Navire / Voyage / Escale — fond pastel, fusionné aussi ──
    subtitle = f"Navire : {navire}     Voyage : {voyage}"
    if escale_info and escale_info.get("date_escale"):
        subtitle += f"     Date d'escale (ETA/ATA) : {escale_info['date_escale']:%d/%m/%Y}"
    ws.append([subtitle])
    subtitle_row = ws.max_row
    ws.merge_cells(start_row=subtitle_row, start_column=1, end_row=subtitle_row, end_column=n_cols)
    sub_cell = ws.cell(row=subtitle_row, column=1)
    sub_cell.font = subtitle_font
    sub_cell.fill = subtitle_fill
    sub_cell.alignment = center
    ws.append([])

    # ── En-tête à 2 niveaux ──
    header_row1 = ws.max_row + 1
    header_row2 = header_row1 + 1
    ws.cell(row=header_row1, column=1, value="POL")
    ws.merge_cells(start_row=header_row1, start_column=1, end_row=header_row2, end_column=1)
    col = 2
    for t in TRANCHES:
        ws.cell(row=header_row1, column=col, value=TRANCHE_LABELS[t])
        ws.merge_cells(start_row=header_row1, start_column=col, end_row=header_row1, end_column=col + 2)
        for j, sub in enumerate(("NOMBRE", "TONNAGE (T)", "VOLUME (M³)")):
            ws.cell(row=header_row2, column=col + j, value=sub)
        col += 3
    for r in (header_row1, header_row2):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

    def _tranche_cols(t):
        i = TRANCHES.index(t)
        return 2 + i * 3, 3 + i * 3, 4 + i * 3  # NOMBRE, TONNAGE, VOLUME

    # ── Une ligne agrégée par POL (zébrage léger, une ligne sur deux) + ligne
    # TOTAL en bas — mêmes chiffres que pivot_pol_tranche, pas de détail
    # conteneur par conteneur ──
    pivot = pivot_pol_tranche(df_classifie)
    row = header_row2
    total_conteneurs = 0
    for i, (_, pr) in enumerate(pivot.iterrows()):
        row += 1
        is_total = pr["POL"] == "TOTAL"
        pol_cell = ws.cell(row=row, column=1, value=pr["POL"])
        pol_cell.font = total_font if is_total else pol_font
        pol_cell.alignment = left
        for t in TRANCHES:
            c_nb, c_tn, c_vol = _tranche_cols(t)
            n = pr[f"{TRANCHE_LABELS[t]} - NOMBRE"]
            if n:
                ws.cell(row=row, column=c_nb, value=n).number_format = "#,##0"
                ws.cell(row=row, column=c_tn, value=pr[f"{TRANCHE_LABELS[t]} - TONNAGE"]).number_format = "#,##0.000"
                ws.cell(row=row, column=c_vol, value=pr[f"{TRANCHE_LABELS[t]} - VOLUME"]).number_format = "#,##0.00"
            if is_total:
                total_conteneurs += n
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.border = border
            if is_total:
                cell.font = total_font
                cell.fill = total_fill
            else:
                if c > 1:
                    cell.font = body_font
                if i % 2 == 1:
                    cell.fill = zebra_fill
            if c > 1:
                cell.alignment = Alignment(horizontal="right", vertical="center")

    if pivot.empty or total_conteneurs == 0:
        ws.cell(row=row + 1, column=1, value="Aucun conteneur classifiable pour cette sélection.").font = body_font

    ws.column_dimensions["A"].width = 24
    for c in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.freeze_panes = f"A{header_row2 + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
